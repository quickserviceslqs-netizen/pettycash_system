"""
Bulk User Import via Excel Template
Allows admins to upload multiple user invitations at once
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db import transaction
from django.core.mail import send_mail
from django.conf import settings
from datetime import timedelta
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd
from accounts.models import User
from accounts.models_device import UserInvitation
from organization.models import Company, Department, Branch
from settings_manager.models import get_setting
from settings_manager.views import log_activity


@login_required
@permission_required('accounts.add_userinvitation', raise_exception=True)
def download_template(request):
    """
    Download Excel template for bulk user import with organization reference
    Supports filtering by company, region, branch, or department
    """
    # Get filter parameters
    company_id = request.GET.get('company')
    region_id = request.GET.get('region')
    branch_id = request.GET.get('branch')
    department_id = request.GET.get('department')
    
    # Build filename based on filters
    from organization.models import Region
    filename_parts = ['user_import']
    filter_context = {}
    
    if department_id:
        dept = Department.objects.filter(id=department_id).select_related('branch__region__company').first()
        if dept:
            filename_parts.append(dept.name.replace(' ', '_'))
            filter_context['department'] = dept
            filter_context['branch'] = dept.branch
            filter_context['region'] = dept.branch.region
            filter_context['company'] = dept.branch.region.company
    elif branch_id:
        branch = Branch.objects.filter(id=branch_id).select_related('region__company').first()
        if branch:
            filename_parts.append(branch.name.replace(' ', '_'))
            filter_context['branch'] = branch
            filter_context['region'] = branch.region
            filter_context['company'] = branch.region.company
    elif region_id:
        region = Region.objects.filter(id=region_id).select_related('company').first()
        if region:
            filename_parts.append(region.name.replace(' ', '_'))
            filter_context['region'] = region
            filter_context['company'] = region.company
    elif company_id:
        company = Company.objects.filter(id=company_id).first()
        if company:
            filename_parts.append(company.name.replace(' ', '_'))
            filter_context['company'] = company
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "User Import"
    
    # Headers with styling
    headers = [
        'email', 'first_name', 'last_name', 'role', 
        'company_name', 'region_name', 'branch_name', 'department_name',
        'assigned_apps'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Get organizations based on filters
    if 'department' in filter_context:
        companies = [filter_context['company']]
        departments = [filter_context['department']]
        branches = [filter_context['branch']]
        regions = [filter_context['region']]
    elif 'branch' in filter_context:
        companies = [filter_context['company']]
        departments = Department.objects.filter(branch=filter_context['branch'])
        branches = [filter_context['branch']]
        regions = [filter_context['region']]
    elif 'region' in filter_context:
        companies = [filter_context['company']]
        branches = Branch.objects.filter(region=filter_context['region'])
        departments = Department.objects.filter(branch__region=filter_context['region'])
        regions = [filter_context['region']]
    elif 'company' in filter_context:
        companies = [filter_context['company']]
        regions = Region.objects.filter(company=filter_context['company'])
        branches = Branch.objects.filter(region__company=filter_context['company'])
        departments = Department.objects.filter(branch__region__company=filter_context['company'])
    else:
        companies = Company.objects.all()[:5]
        departments = Department.objects.all()[:5]
        branches = Branch.objects.all()[:5]
        regions = Region.objects.all()[:5]
    
    # Add example rows
    company_example = companies[0].name if companies else 'Quick Services LQS'
    dept_example = departments[0].name if departments else 'Finance'
    branch_example = branches[0].name if branches else 'Head Office'
    region_example = regions[0].name if regions else 'East Africa'
    
    ws.append([
        'amos.cheloti@example.com', 'Amos', 'Cheloti', 'REQUESTER',
        company_example, region_example, branch_example, dept_example, 'treasury,workflow'
    ])
    ws.append([
        'jane.doe@example.com', 'Jane', 'Doe', 'APPROVER',
        company_example, region_example, branch_example, dept_example, 'treasury,workflow,reports'
    ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30  # email
    ws.column_dimensions['B'].width = 15  # first_name
    ws.column_dimensions['C'].width = 15  # last_name
    ws.column_dimensions['D'].width = 12  # role
    ws.column_dimensions['E'].width = 20  # company
    ws.column_dimensions['F'].width = 15  # region
    ws.column_dimensions['G'].width = 20  # branch
    ws.column_dimensions['H'].width = 20  # department
    ws.column_dimensions['I'].width = 30  # apps
    
    # Create Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2['A1'] = "Bulk User Import Instructions"
    ws2['A1'].font = Font(bold=True, size=14)
    
    instructions = [
        "",
        "FIELD REQUIREMENTS:",
        "• email: Must be valid and unique (user@company.com)",
        "• first_name: User's first name (e.g., Amos)",
        "• last_name: User's last name (e.g., Cheloti)",
        "• role: REQUESTER | APPROVER | FINANCE | ADMIN",
        "• company_name: Must EXACTLY match existing company",
        "• region_name: Region name (optional)",
        "• branch_name: Must EXACTLY match existing branch",
        "• department_name: Must EXACTLY match existing department",
        "• assigned_apps: Comma-separated (treasury,workflow,reports,settings)",
        "",
        "USERNAME FORMAT:",
        "• Auto-generated as: FirstInitial.LastName",
        "• Example: Amos Cheloti becomes A.Cheloti",
        "• Duplicates get number suffix: A.Cheloti1, A.Cheloti2",
        "",
        "PASSWORD:",
        "• Users will receive email invitation to set their own password",
        "",
        "AVAILABLE ORGANIZATIONS:",
        ""
    ]
    
    # Add companies
    instructions.append("COMPANIES:")
    all_companies = companies if filter_context else Company.objects.all().order_by('name')
    for company in all_companies:
        instructions.append(f"  • {company.name}")
    if not all_companies:
        instructions.append("  (No companies found - create companies first)")
    instructions.append("")
    
    # Add regions
    instructions.append("REGIONS:")
    if filter_context:
        all_regions = regions if 'region' in filter_context or 'branch' in filter_context or 'department' in filter_context else Region.objects.filter(company__in=companies)
    else:
        all_regions = Region.objects.all().order_by('name')[:20]
    for region in all_regions:
        company_info = f' ({region.company.name})' if hasattr(region, 'company') and region.company else ''
        instructions.append(f"  • {region.name}{company_info}")
    if not all_regions:
        instructions.append("  (No regions found)")
    instructions.append("")
    
    # Add branches
    instructions.append("BRANCHES:")
    all_branches = branches if filter_context else Branch.objects.all().order_by('name')[:20]
    for branch in all_branches:
        region_info = f' ({branch.region.name})' if hasattr(branch, 'region') and branch.region else ''
        instructions.append(f"  • {branch.name}{region_info}")
    if not all_branches:
        instructions.append("  (No branches found)")
    instructions.append("")
    
    # Add departments
    instructions.append("DEPARTMENTS:")
    all_departments = departments if filter_context else Department.objects.all().order_by('name')[:20]
    for dept in all_departments:
        branch_info = f' ({dept.branch.name})' if hasattr(dept, 'branch') and dept.branch else ''
        instructions.append(f"  • {dept.name}{branch_info}")
    if not all_departments:
        instructions.append("  (No departments found)")
    instructions.append("")
    
    # Add roles
    instructions.extend([
        "AVAILABLE ROLES:",
        "  • REQUESTER - Can create requests",
        "  • APPROVER - Can approve requests",
        "  • FINANCE - Finance team member",
        "  • ADMIN - System administrator",
        "",
        "AVAILABLE APPS:",
        "  • treasury - Treasury management",
        "  • workflow - Workflow approval",
        "  • reports - Reports and analytics",
        "  • settings - System settings",
        "  • transactions - Transaction management",
        "  • organization - Organization management"
    ])
    
    # Write instructions to sheet
    for row_num, instruction in enumerate(instructions, 2):
        ws2[f'A{row_num}'] = instruction
    
    ws2.column_dimensions['A'].width = 70
    
    # Prepare response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{"_".join(filename_parts)}_template.xlsx"'
    wb.save(response)
    
    return response


@login_required
@permission_required('accounts.add_userinvitation', raise_exception=True)
def bulk_import(request):
    """
    Upload and process CSV or Excel file for bulk user invitations
    """
    if request.method == 'POST' and request.FILES.get('csv_file'):
        uploaded_file = request.FILES['csv_file']
        
        # Validate file type (support both CSV and Excel)
        is_excel = uploaded_file.name.endswith(('.xlsx', '.xls'))
        is_csv = uploaded_file.name.endswith('.csv')
        
        if not (is_excel or is_csv):
            messages.error(request, 'Please upload a CSV or Excel file (.csv, .xlsx, .xls)')
            return redirect('accounts:bulk_import')
        
        # Read and decode file
        try:
            if is_excel:
                # Read Excel file using pandas
                df = pd.read_excel(uploaded_file)
                # Convert to list of dicts
                data_rows = df.to_dict('records')
            else:
                # Read CSV file
                decoded_file = uploaded_file.read().decode('utf-8')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                data_rows = list(reader)
            
            success_count = 0
            error_count = 0
            errors = []
            
            with transaction.atomic():
                for row_num, row in enumerate(data_rows, start=2):  # Start at 2 (after header)
                    try:
                        # Convert pandas values to strings and handle NaN
                        if is_excel:
                            row = {k: (str(v) if pd.notna(v) else '') for k, v in row.items()}
                        
                        # Skip empty rows or instruction rows
                        if not row.get('email') or str(row.get('email', '')).startswith('INSTRUCTIONS'):
                            continue
                        
                        # Validate required fields
                        email = row['email'].strip()
                        first_name = row['first_name'].strip()
                        last_name = row['last_name'].strip()
                        role = row['role'].strip().upper()
                        
                        if not all([email, first_name, last_name, role]):
                            errors.append(f"Row {row_num}: Missing required fields")
                            error_count += 1
                            continue
                        
                        # Validate role
                        valid_roles = [choice[0] for choice in User.ROLE_CHOICES]
                        if role not in valid_roles:
                            errors.append(f"Row {row_num}: Invalid role '{role}'. Must be one of: {', '.join(valid_roles)}")
                            error_count += 1
                            continue
                        
                        # Check if user or invitation already exists
                        if User.objects.filter(email=email).exists():
                            errors.append(f"Row {row_num}: User with email {email} already exists")
                            error_count += 1
                            continue
                        
                        if UserInvitation.objects.filter(email=email, status='pending').exists():
                            errors.append(f"Row {row_num}: Pending invitation already exists for {email}")
                            error_count += 1
                            continue
                        
                        # Get organizational entities
                        company = None
                        department = None
                        branch = None
                        region_name = row.get('region_name', '').strip() if row.get('region_name') else None
                        
                        if row.get('company_name'):
                            company_name = row['company_name'].strip()
                            try:
                                company = Company.objects.get(name=company_name)
                            except Company.DoesNotExist:
                                # Try case-insensitive match
                                companies = Company.objects.filter(name__iexact=company_name)
                                if companies.exists():
                                    company = companies.first()
                                else:
                                    errors.append(f"Row {row_num}: Company '{company_name}' not found. Check exact spelling.")
                                    error_count += 1
                                    continue
                        
                        if row.get('department_name'):
                            dept_name = row['department_name'].strip()
                            try:
                                department = Department.objects.get(name=dept_name)
                            except Department.DoesNotExist:
                                # Try case-insensitive match
                                departments = Department.objects.filter(name__iexact=dept_name)
                                if departments.exists():
                                    department = departments.first()
                                else:
                                    errors.append(f"Row {row_num}: Department '{dept_name}' not found. Check exact spelling.")
                                    error_count += 1
                                    continue
                        
                        if row.get('branch_name'):
                            branch_name = row['branch_name'].strip()
                            try:
                                # If region specified, filter by region too
                                if region_name:
                                    branch = Branch.objects.get(name=branch_name, region__name=region_name)
                                else:
                                    branch = Branch.objects.get(name=branch_name)
                            except Branch.DoesNotExist:
                                # Try case-insensitive match
                                if region_name:
                                    branches = Branch.objects.filter(name__iexact=branch_name, region__name__iexact=region_name)
                                else:
                                    branches = Branch.objects.filter(name__iexact=branch_name)
                                
                                if branches.exists():
                                    branch = branches.first()
                                else:
                                    region_hint = f" in region '{region_name}'" if region_name else ""
                                    errors.append(f"Row {row_num}: Branch '{branch_name}'{region_hint} not found. Check exact spelling.")
                                    error_count += 1
                                    continue
                            except Branch.MultipleObjectsReturned:
                                errors.append(f"Row {row_num}: Multiple branches named '{branch_name}' found. Please specify region_name.")
                                error_count += 1
                                continue
                        
                        # Parse assigned apps
                        assigned_apps_list = []
                        if row.get('assigned_apps'):
                            app_names = [app.strip() for app in row['assigned_apps'].split(',')]
                            assigned_apps_list = app_names
                        
                        # Get invitation expiry
                        expiry_days = int(get_setting('INVITATION_EXPIRY_DAYS', '7'))
                        expires_at = timezone.now() + timedelta(days=expiry_days)
                        
                        # Create invitation
                        invitation = UserInvitation.objects.create(
                            email=email,
                            first_name=first_name,
                            last_name=last_name,
                            role=role,
                            company=company,
                            department=department,
                            branch=branch,
                            invited_by=request.user,
                            expires_at=expires_at,
                            assigned_apps=assigned_apps_list
                        )
                        
                        # Send invitation email
                        try:
                            invitation_url = request.build_absolute_uri(
                                f'/accounts/signup/{invitation.token}/'
                            )
                            
                            # Generate username preview for email
                            first_initial = first_name[0].upper() if first_name else 'U'
                            clean_last = last_name.replace(' ', '').replace('-', '').replace("'", '')
                            username_preview = f"{first_initial}.{clean_last}"
                            
                            send_mail(
                                subject=f'Invitation to join {company.name if company else "the system"}',
                                message=f'''
Hello {first_name},

You've been invited to join as a {invitation.get_role_display()}.

Click here to complete your registration:
{invitation_url}

Your username will be auto-generated as: {username_preview}
You will set your password during registration.

This invitation expires on {expires_at.strftime("%B %d, %Y at %I:%M %p")}.

Best regards,
{request.user.get_full_name() or request.user.username}
                                ''',
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[email],
                                fail_silently=False,
                            )
                        except Exception as e:
                            errors.append(f"Row {row_num}: Email sent failed for {email}: {str(e)}")
                        
                        success_count += 1
                        
                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(
                    request, 
                    f'✅ Successfully created {success_count} invitation(s)'
                )
                log_activity(
                    request.user,
                    'USER_BULK_IMPORT',
                    f'Bulk imported {success_count} user invitations'
                )
            
            if error_count > 0:
                messages.warning(
                    request,
                    f'⚠️ {error_count} row(s) failed. See details below.'
                )
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                
                if len(errors) > 10:
                    messages.info(request, f'... and {len(errors) - 10} more errors')
            
            return redirect('accounts:manage_invitations')
            
        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')
            return redirect('accounts:bulk_import')
    
    # GET request - show upload form
    from organization.models import Region
    context = {
        'companies': Company.objects.all().order_by('name'),
        'regions': Region.objects.select_related('company').all().order_by('company__name', 'name'),
        'departments': Department.objects.select_related('branch__region__company').all().order_by('branch__region__company__name', 'name'),
        'branches': Branch.objects.select_related('region__company').all().order_by('region__company__name', 'region__name', 'name'),
        'roles': User.ROLE_CHOICES,
    }
    return render(request, 'accounts/bulk_import.html', context)
