"""
App Management Views
Provides UI for managing application modules and user app assignments.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Count, Q
from django.http import JsonResponse, HttpResponse

from accounts.models import App, User
import openpyxl
from openpyxl.styles import Font, PatternFill
import pandas as pd


def is_admin_user(user):
    """Check if user is superuser or has admin role"""
    return user.is_superuser or user.role == "admin"


@login_required
@user_passes_test(is_admin_user)
def manage_apps(request):
    """
    Dashboard for viewing and managing all applications.
    Shows app details, active status, and user assignments.
    """
    # Get all apps with user counts
    apps = App.objects.annotate(user_count=Count("users")).order_by("display_name")

    # Apply filters
    status_filter = request.GET.get("status", "")
    search_query = request.GET.get("search", "").strip()

    if status_filter == "active":
        apps = apps.filter(is_active=True)
    elif status_filter == "inactive":
        apps = apps.filter(is_active=False)

    if search_query:
        apps = apps.filter(
            Q(display_name__icontains=search_query)
            | Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
        )

    # Statistics
    stats = {
        "total": App.objects.count(),
        "active": App.objects.filter(is_active=True).count(),
        "inactive": App.objects.filter(is_active=False).count(),
        "total_assignments": User.objects.filter(assigned_apps__isnull=False)
        .distinct()
        .count(),
    }

    context = {
        "apps": apps,
        "stats": stats,
        "filters": {
            "status": status_filter,
            "search": search_query,
        },
        "app_choices": App.APP_CHOICES,
    }

    return render(request, "accounts/manage_apps.html", context)


@login_required
@user_passes_test(is_admin_user)
def create_app(request):
    """
    Create a new application module.
    """
    if request.method == "POST":
        name = request.POST.get("name")
        display_name = request.POST.get("display_name")
        url = request.POST.get("url")
        description = request.POST.get("description", "")
        is_active = request.POST.get("is_active") == "on"

        # Validate
        if not name:
            messages.error(request, "App name is required.")
            return redirect("accounts:create_app")

        # Check if app already exists
        if App.objects.filter(name=name).exists():
            messages.error(request, f"App '{name}' already exists.")
            return redirect("accounts:create_app")

        # Create app
        app = App.objects.create(
            name=name,
            display_name=display_name or name.capitalize(),
            url=url or f"/{name}/",
            description=description,
            is_active=is_active,
        )

        messages.success(request, f"App '{app.display_name}' created successfully.")
        return redirect("accounts:manage_apps")

    context = {
        "app_choices": App.APP_CHOICES,
    }
    return render(request, "accounts/create_app.html", context)


@login_required
@user_passes_test(is_admin_user)
def edit_app(request, app_id):
    """
    Edit an existing application module.
    """
    app = get_object_or_404(App, id=app_id)

    if request.method == "POST":
        app.display_name = request.POST.get("display_name", app.display_name)
        app.url = request.POST.get("url", app.url)
        app.description = request.POST.get("description", "")
        app.is_active = request.POST.get("is_active") == "on"

        app.save()

        messages.success(request, f"App '{app.display_name}' updated successfully.")
        return redirect("accounts:manage_apps")

    # Get users assigned to this app
    assigned_users = (
        User.objects.filter(assigned_apps=app)
        .select_related("company", "department", "branch")
        .order_by("first_name", "last_name")
    )

    context = {
        "app": app,
        "assigned_users": assigned_users,
    }
    return render(request, "accounts/edit_app.html", context)


@login_required
@user_passes_test(is_admin_user)
def toggle_app_status(request, app_id):
    """
    Toggle app active/inactive status via AJAX.
    """
    if request.method == "POST":
        app = get_object_or_404(App, id=app_id)
        app.is_active = not app.is_active
        app.save()

        return JsonResponse(
            {
                "success": True,
                "is_active": app.is_active,
                "message": f"App '{app.display_name}' is now {'active' if app.is_active else 'inactive'}.",
            }
        )

    return JsonResponse({"success": False, "message": "POST request required"})


@login_required
@user_passes_test(is_admin_user)
def assign_apps_to_user(request, user_id):
    """
    Assign/unassign apps to a specific user.
    """
    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        app_ids = request.POST.getlist("apps")
        apps = App.objects.filter(id__in=app_ids)

        user.assigned_apps.set(apps)

        messages.success(
            request,
            f"Updated app assignments for {user.get_full_name()}. Assigned {apps.count()} apps.",
        )
        return redirect("accounts:manage_users")

    # Get all apps and user's current assignments
    all_apps = App.objects.filter(is_active=True).order_by("display_name")
    user_app_ids = list(user.assigned_apps.values_list("id", flat=True))

    context = {
        "user": user,
        "all_apps": all_apps,
        "user_app_ids": user_app_ids,
    }
    return render(request, "accounts/assign_apps.html", context)


@login_required
@user_passes_test(is_admin_user)
def bulk_assign_apps(request):
    """
    Bulk assign apps to multiple users via Excel upload.
    """
    if request.method == "POST" and request.FILES.get("file"):
        try:
            excel_file = request.FILES["file"]
            df = pd.read_excel(excel_file)

            # Validate columns
            required_columns = ["email", "app_name"]
            if not all(col in df.columns for col in required_columns):
                messages.error(
                    request,
                    f"Excel file must contain columns: {', '.join(required_columns)}",
                )
                return redirect("accounts:bulk_assign_apps")

            results = {"total": len(df), "successful": 0, "failed": 0, "errors": []}

            for index, row in df.iterrows():
                try:
                    email = str(row["email"]).strip()
                    app_name = str(row["app_name"]).strip()

                    # Find user
                    try:
                        user = User.objects.get(email=email)
                    except User.DoesNotExist:
                        results["errors"].append(
                            f"Row {index + 2}: User not found: {email}"
                        )
                        results["failed"] += 1
                        continue

                    # Find app
                    try:
                        app = App.objects.get(name=app_name)
                    except App.DoesNotExist:
                        results["errors"].append(
                            f"Row {index + 2}: App not found: {app_name}"
                        )
                        results["failed"] += 1
                        continue

                    # Assign app to user
                    user.assigned_apps.add(app)
                    results["successful"] += 1

                except Exception as e:
                    results["errors"].append(f"Row {index + 2}: {str(e)}")
                    results["failed"] += 1

            messages.success(
                request,
                f"Processed {results['total']} rows: {results['successful']} successful, {results['failed']} failed.",
            )

            context = {
                "results": results,
            }
            return render(request, "accounts/bulk_assign_apps.html", context)

        except Exception as e:
            messages.error(request, f"Error processing file: {str(e)}")
            return redirect("accounts:bulk_assign_apps")

    return render(request, "accounts/bulk_assign_apps.html")


@login_required
@user_passes_test(is_admin_user)
def download_bulk_assign_template(request):
    """
    Download Excel template for bulk app assignment.
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk App Assignment"

    # Headers
    headers = ["email", "app_name"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # Add example rows
    ws.append(["user@example.com", "Treasury"])
    ws.append(["user2@example.com", "Transactions"])

    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "Bulk App Assignment Instructions"
    ws2["A1"].font = Font(bold=True, size=14)

    instructions = [
        "",
        "1. Fill in the 'Bulk App Assignment' sheet with user emails and app names",
        "2. Available app names:",
    ]

    # Get all active apps
    apps = App.objects.filter(is_active=True).order_by("display_name")
    for app in apps:
        instructions.append(f"   - {app.name}")

    instructions.extend(
        [
            "",
            "3. Save the file and upload it in the Bulk Assign Apps page",
            "4. Each row represents one user-app assignment",
            "",
            "Note: Make sure the email addresses and app names match exactly",
        ]
    )

    for row_num, instruction in enumerate(instructions, 2):
        ws2[f"A{row_num}"] = instruction

    # Adjust column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws2.column_dimensions["A"].width = 60

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        "attachment; filename=bulk_app_assignment_template.xlsx"
    )
    wb.save(response)

    return response


@login_required
@user_passes_test(is_admin_user)
def app_users(request, app_id):
    """
    View all users assigned to a specific app.
    """
    app = get_object_or_404(App, id=app_id)

    users = (
        User.objects.filter(assigned_apps=app)
        .select_related("company", "region", "branch", "department")
        .order_by("first_name", "last_name")
    )

    # Apply filters
    role_filter = request.GET.get("role", "")
    company_filter = request.GET.get("company", "")
    search_query = request.GET.get("search", "").strip()

    if role_filter:
        users = users.filter(role=role_filter)

    if company_filter:
        users = users.filter(company_id=company_filter)

    if search_query:
        users = users.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(username__icontains=search_query)
        )

    context = {
        "app": app,
        "users": users,
        "total_users": users.count(),
        "filters": {
            "role": role_filter,
            "company": company_filter,
            "search": search_query,
        },
    }

    return render(request, "accounts/app_users.html", context)
