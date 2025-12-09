import csv
from datetime import timedelta

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import (
    Avg,
    Count,
    DurationField,
    Exists,
    ExpressionWrapper,
    F,
    Min,
    OuterRef,
    Q,
    Subquery,
    Sum,
)
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook

from accounts.permissions import require_app_access, require_permission
from transactions.models import ApprovalTrail, Requisition
from treasury.models import LedgerEntry, Payment, PaymentExecution, TreasuryFund

User = get_user_model()


def is_admin_user(user):
    # Deprecated: replaced by app access + permission checks
    return bool(user and user.is_authenticated)


@require_app_access("reports")
@require_permission("view_reports_dashboard", app_label="reports")
def reports_dashboard(request):
    """
    Main reports dashboard with overview metrics and quick access to detailed reports.
    Shows key statistics across transactions, treasury, and approvals.
    """
    # Get date range from request (default: last 30 days)
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)

    # Transaction Statistics
    requisitions = Requisition.objects.filter(created_at__gte=start_date)
    total_requisitions = requisitions.count()
    total_amount = requisitions.aggregate(Sum("amount"))["amount__sum"] or 0

    # Status breakdown
    status_breakdown = (
        requisitions.values("status")
        .annotate(count=Count("transaction_id"))
        .order_by("-count")
    )

    # Pending requisitions (not paid, reviewed, or rejected)
    pending_count = requisitions.exclude(
        status__in=["paid", "reviewed", "rejected"]
    ).count()
    paid_count = requisitions.filter(status="paid").count()
    rejected_count = requisitions.filter(status="rejected").count()

    # Treasury Statistics
    total_treasury_balance = (
        TreasuryFund.objects.aggregate(Sum("current_balance"))["current_balance__sum"]
        or 0
    )
    funds_below_reorder = TreasuryFund.objects.filter(
        current_balance__lt=F("reorder_level")
    ).count()

    # Payment Statistics
    payments = Payment.objects.filter(requisition__created_at__gte=start_date)
    successful_payments = payments.filter(status="success").count()
    failed_payments = payments.filter(status="failed").count()
    total_paid_amount = (
        payments.filter(status="success").aggregate(Sum("amount"))["amount__sum"] or 0
    )

    # User Activity
    active_users = requisitions.values("requested_by").distinct().count()

    # Approval Metrics
    approval_trails = ApprovalTrail.objects.filter(timestamp__gte=start_date)
    total_approvals = approval_trails.filter(action="approved").count()
    total_rejections = approval_trails.filter(action="rejected").count()
    avg_approval_time = approval_trails.filter(action="approved").aggregate(
        avg_time=Avg(F("timestamp") - F("requisition__created_at"))
    )["avg_time"]

    # Top requesters
    top_requesters = (
        requisitions.values(
            "requested_by__username",
            "requested_by__first_name",
            "requested_by__last_name",
        )
        .annotate(total=Count("transaction_id"), amount=Sum("amount"))
        .order_by("-total")[:5]
    )

    # Top approvers
    top_approvers = (
        approval_trails.filter(action="approved")
        .values("user__username")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    context = {
        "days": days,
        "start_date": start_date,
        # Transaction metrics
        "total_requisitions": total_requisitions,
        "total_amount": total_amount,
        "pending_count": pending_count,
        "paid_count": paid_count,
        "rejected_count": rejected_count,
        "status_breakdown": status_breakdown,
        # Treasury metrics
        "total_treasury_balance": total_treasury_balance,
        "funds_below_reorder": funds_below_reorder,
        # Payment metrics
        "successful_payments": successful_payments,
        "failed_payments": failed_payments,
        "total_paid_amount": total_paid_amount,
        # Activity metrics
        "active_users": active_users,
        "total_approvals": total_approvals,
        "total_rejections": total_rejections,
        "avg_approval_time": avg_approval_time,
        # Top users
        "top_requesters": top_requesters,
        "top_approvers": top_approvers,
    }

    return render(request, "reports/dashboard.html", context)


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def transaction_report(request):
    """
    Detailed transaction/requisition report with filters and export capability.
    """
    # Get filter parameters
    days = int(request.GET.get("days", 30))
    status = request.GET.get("status", "")
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    min_amount = request.GET.get("min_amount", "")
    max_amount = request.GET.get("max_amount", "")

    start_date = timezone.now() - timedelta(days=days)

    # Base queryset
    requisitions = (
        Requisition.objects.filter(created_at__gte=start_date)
        .select_related("requested_by", "branch", "department", "cost_center")
        .order_by("-created_at")
    )

    # Apply filters
    if status:
        requisitions = requisitions.filter(status=status)
    if branch_id:
        requisitions = requisitions.filter(branch_id=branch_id)
    if department_id:
        requisitions = requisitions.filter(department_id=department_id)
    if min_amount:
        requisitions = requisitions.filter(amount__gte=min_amount)
    if max_amount:
        requisitions = requisitions.filter(amount__lte=max_amount)

    # Statistics
    total_count = requisitions.count()
    total_amount = requisitions.aggregate(Sum("amount"))["amount__sum"] or 0
    avg_amount = requisitions.aggregate(Avg("amount"))["amount__avg"] or 0

    # Pagination
    from django.core.paginator import Paginator

    try:
        page_size = int(request.GET.get("page_size", 25))
    except (TypeError, ValueError):
        page_size = 25
    page_size = max(5, min(page_size, 200))
    page_number = request.GET.get("page", 1)
    paginator = Paginator(requisitions, page_size)
    page_obj = paginator.get_page(page_number)

    # Get unique branches and departments for filters
    from organization.models import Branch, Department

    branches = Branch.objects.all().order_by("name")
    departments = Department.objects.all().order_by("name")

    context = {
        "requisitions": page_obj,
        "total_count": total_count,
        "total_amount": total_amount,
        "avg_amount": avg_amount,
        "days": days,
        "page_obj": page_obj,
        "page_size": page_size,
        "selected_status": status,
        "selected_branch": branch_id,
        "selected_department": department_id,
        "min_amount": min_amount,
        "max_amount": max_amount,
        "branches": branches,
        "departments": departments,
        "status_choices": Requisition.STATUS_CHOICES,
    }

    return render(request, "reports/transaction_report.html", context)


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def transaction_report_export_csv(request):
    """Server-side CSV export for Transaction Report honoring filters."""
    # Filters (same as transaction_report)
    days = int(request.GET.get("days", 30))
    status = request.GET.get("status", "")
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    min_amount = request.GET.get("min_amount", "")
    max_amount = request.GET.get("max_amount", "")

    start_date = timezone.now() - timedelta(days=days)
    qs = (
        Requisition.objects.filter(created_at__gte=start_date)
        .select_related("requested_by", "branch", "department", "cost_center")
        .order_by("-created_at")
    )

    if status:
        qs = qs.filter(status=status)
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if min_amount:
        qs = qs.filter(amount__gte=min_amount)
    if max_amount:
        qs = qs.filter(amount__lte=max_amount)

    # Build CSV
    response = HttpResponse(content_type="text/csv")
    filename = f"transaction_report_{timezone.now().date()}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    writer.writerow(
        [
            "Transaction ID",
            "Requested By",
            "Branch",
            "Department",
            "Amount",
            "Status",
            "Created",
            "Purpose",
        ]
    )
    for r in qs.iterator():
        writer.writerow(
            [
                r.transaction_id,
                getattr(r.requested_by, "username", ""),
                getattr(r.branch, "name", "") if r.branch else "",
                getattr(r.department, "name", "") if r.department else "",
                f"{r.amount}",
                r.get_status_display(),
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                (r.purpose or "").replace("\n", " ").strip(),
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def transaction_report_export_xlsx(request):
    """Server-side Excel export for Transaction Report honoring filters."""
    # Filters (same as transaction_report)
    days = int(request.GET.get("days", 30))
    status = request.GET.get("status", "")
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    min_amount = request.GET.get("min_amount", "")
    max_amount = request.GET.get("max_amount", "")

    start_date = timezone.now() - timedelta(days=days)
    qs = (
        Requisition.objects.filter(created_at__gte=start_date)
        .select_related("requested_by", "branch", "department", "cost_center")
        .order_by("-created_at")
    )

    if status:
        qs = qs.filter(status=status)
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    if department_id:
        qs = qs.filter(department_id=department_id)
    if min_amount:
        qs = qs.filter(amount__gte=min_amount)
    if max_amount:
        qs = qs.filter(amount__lte=max_amount)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    ws.append(
        [
            "Transaction ID",
            "Requested By",
            "Branch",
            "Department",
            "Amount",
            "Status",
            "Created",
            "Purpose",
        ]
    )
    for r in qs.iterator():
        ws.append(
            [
                r.transaction_id,
                getattr(r.requested_by, "username", ""),
                getattr(r.branch, "name", "") if r.branch else "",
                getattr(r.department, "name", "") if r.department else "",
                float(r.amount),
                r.get_status_display(),
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                (r.purpose or "").replace("\n", " ").strip(),
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"transaction_report_{timezone.now().date()}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_report(request):
    """
    Treasury fund balances and payment activity report.
    """
    # Get all treasury funds with annotations
    funds_qs = (
        TreasuryFund.objects.select_related("company", "region", "branch")
        .annotate(needs_replenishment=Q(current_balance__lt=F("reorder_level")))
        .order_by("company", "region", "branch")
    )

    # Get date range
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)

    # Payment statistics by method
    payment_stats = (
        Payment.objects.filter(requisition__created_at__gte=start_date)
        .values("method")
        .annotate(count=Count("payment_id"), total_amount=Sum("amount"))
        .order_by("-total_amount")
    )

    # Payment statistics by status
    status_stats = (
        Payment.objects.filter(requisition__created_at__gte=start_date)
        .values("status")
        .annotate(count=Count("payment_id"), total_amount=Sum("amount"))
        .order_by("-count")
    )

    # Overall statistics
    total_balance = (
        funds_qs.aggregate(Sum("current_balance"))["current_balance__sum"] or 0
    )
    funds_needing_replenishment = funds_qs.filter(
        current_balance__lt=F("reorder_level")
    ).count()

    # Pagination
    page_size = max(5, min(200, int(request.GET.get("page_size", 25))))
    paginator = Paginator(funds_qs, page_size)
    page_number = request.GET.get("page", 1)
    funds = paginator.get_page(page_number)

    context = {
        "funds": funds,
        "page_size": page_size,
        "total_balance": total_balance,
        "funds_needing_replenishment": funds_needing_replenishment,
        "payment_stats": payment_stats,
        "status_stats": status_stats,
        "days": days,
    }

    return render(request, "reports/treasury_report.html", context)


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_fund_detail(request, fund_id):
    """Drilldown: show ledger movements and executed payments for a fund."""
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)

    fund = TreasuryFund.objects.select_related(
        "company", "region", "branch", "department"
    ).get(fund_id=fund_id)

    # Ledger entries for this fund in date range
    ledger_qs = (
        LedgerEntry.objects.filter(treasury_fund=fund, created_at__gte=start_date)
        .select_related("payment_execution")
        .order_by("-created_at")
    )

    # Payments associated via payment execution -> ledger entry
    payments_qs = (
        Payment.objects.filter(
            execution__ledger_entries__treasury_fund=fund, created_at__gte=start_date
        )
        .select_related("requisition", "executor", "created_by")
        .distinct()
        .order_by("-created_at")
    )

    # Pagination for ledger
    from django.core.paginator import Paginator

    try:
        ledger_page_size = int(request.GET.get("ledger_page_size", 25))
    except (TypeError, ValueError):
        ledger_page_size = 25
    ledger_page_size = max(5, min(ledger_page_size, 200))
    ledger_page = request.GET.get("ledger_page", 1)
    ledger_paginator = Paginator(ledger_qs, ledger_page_size)
    ledger_page_obj = ledger_paginator.get_page(ledger_page)

    # Pagination for payments
    try:
        pay_page_size = int(request.GET.get("pay_page_size", 25))
    except (TypeError, ValueError):
        pay_page_size = 25
    pay_page_size = max(5, min(pay_page_size, 200))
    pay_page = request.GET.get("pay_page", 1)
    pay_paginator = Paginator(payments_qs, pay_page_size)
    pay_page_obj = pay_paginator.get_page(pay_page)

    # Totals
    debits = (
        ledger_qs.filter(entry_type="debit").aggregate(total=Sum("amount"))["total"]
        or 0
    )
    credits = (
        ledger_qs.filter(entry_type="credit").aggregate(total=Sum("amount"))["total"]
        or 0
    )

    context = {
        "fund": fund,
        "days": days,
        "ledger_page_obj": ledger_page_obj,
        "ledger_page_size": ledger_page_size,
        "pay_page_obj": pay_page_obj,
        "pay_page_size": pay_page_size,
        "debits": debits,
        "credits": credits,
    }
    return render(request, "reports/treasury_fund_detail.html", context)


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_fund_ledger_export_csv(request, fund_id):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    fund = TreasuryFund.objects.get(fund_id=fund_id)
    qs = (
        LedgerEntry.objects.filter(treasury_fund=fund, created_at__gte=start_date)
        .select_related("payment_execution")
        .order_by("-created_at")
    )
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="fund_ledger_{fund_id}_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(["Timestamp", "Type", "Amount", "Description", "Payment Ref"])
    for e in qs.iterator():
        writer.writerow(
            [
                e.created_at.strftime("%Y-%m-%d %H:%M"),
                e.entry_type,
                f"{e.amount}",
                (e.description or "").replace("\n", " ").strip(),
                getattr(e.payment_execution, "gateway_reference", ""),
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_fund_ledger_export_xlsx(request, fund_id):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    fund = TreasuryFund.objects.get(fund_id=fund_id)
    qs = (
        LedgerEntry.objects.filter(treasury_fund=fund, created_at__gte=start_date)
        .select_related("payment_execution")
        .order_by("-created_at")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Ledger"
    ws.append(["Timestamp", "Type", "Amount", "Description", "Payment Ref"])
    for e in qs.iterator():
        ws.append(
            [
                e.created_at.strftime("%Y-%m-%d %H:%M"),
                e.entry_type,
                float(e.amount),
                (e.description or "").replace("\n", " ").strip(),
                getattr(e.payment_execution, "gateway_reference", ""),
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="fund_ledger_{fund_id}_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_fund_payments_export_csv(request, fund_id):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    fund = TreasuryFund.objects.get(fund_id=fund_id)
    qs = (
        Payment.objects.filter(
            execution__ledger_entries__treasury_fund=fund, created_at__gte=start_date
        )
        .select_related("requisition", "executor", "created_by")
        .distinct()
        .order_by("-created_at")
    )
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="fund_payments_{fund_id}_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(
        [
            "Created",
            "Payment ID",
            "Amount",
            "Method",
            "Status",
            "Destination",
            "Requisition",
            "Executor",
        ]
    )
    for p in qs.iterator():
        writer.writerow(
            [
                p.created_at.strftime("%Y-%m-%d %H:%M"),
                p.payment_id,
                f"{p.amount}",
                p.method,
                p.status,
                (p.destination or ""),
                getattr(p.requisition, "transaction_id", ""),
                getattr(p.executor, "username", ""),
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def treasury_fund_payments_export_xlsx(request, fund_id):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    fund = TreasuryFund.objects.get(fund_id=fund_id)
    qs = (
        Payment.objects.filter(
            execution__ledger_entries__treasury_fund=fund, created_at__gte=start_date
        )
        .select_related("requisition", "executor", "created_by")
        .distinct()
        .order_by("-created_at")
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Fund Payments"
    ws.append(
        [
            "Created",
            "Payment ID",
            "Amount",
            "Method",
            "Status",
            "Destination",
            "Requisition",
            "Executor",
        ]
    )
    for p in qs.iterator():
        ws.append(
            [
                p.created_at.strftime("%Y-%m-%d %H:%M"),
                str(p.payment_id),
                float(p.amount),
                p.method,
                p.status,
                (p.destination or ""),
                getattr(p.requisition, "transaction_id", ""),
                getattr(p.executor, "username", ""),
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="fund_payments_{fund_id}_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def approval_report(request):
    """
    Approval workflow analytics: approval times, bottlenecks, approver performance.
    """
    # Get date range
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)

    # Approval trails with related data
    approval_trails = (
        ApprovalTrail.objects.filter(timestamp__gte=start_date)
        .select_related("user", "requisition")
        .order_by("-timestamp")
    )

    # Approver performance - exclude non-approver actions like 'changes_submitted', 'paid', etc.
    approver_actions = approval_trails.filter(
        action__in=[
            "approved",
            "rejected",
            "validated",
            "reviewed",
            "changes_requested",
            "urgency_confirmed",
            "reverted_to_normal",
        ]
    )
    approver_stats = (
        approver_actions.values("user__username", "user__id", "user__role")
        .annotate(
            total_actions=Count("id"),
            approvals=Count("id", filter=Q(action="approved")),
            rejections=Count("id", filter=Q(action="rejected")),
            change_requests=Count("id", filter=Q(action="changes_requested")),
        )
        .exclude(user__role="staff")
        .order_by("-total_actions")
    )

    # Action breakdown
    action_breakdown = (
        approval_trails.values("action").annotate(count=Count("id")).order_by("-count")
    )

    # SLA: average time to first approval (hours)
    first_approved_subq = (
        ApprovalTrail.objects.filter(requisition_id=OuterRef("pk"), action="approved")
        .order_by()
        .values("requisition_id")
        .annotate(first_at=Min("timestamp"))
        .values("first_at")[:1]
    )
    req_with_first = (
        Requisition.objects.filter(created_at__gte=start_date)
        .annotate(first_approved_at=Subquery(first_approved_subq))
        .filter(first_approved_at__isnull=False)
        .annotate(
            time_to_first=ExpressionWrapper(
                F("first_approved_at") - F("created_at"), output_field=DurationField()
            )
        )
    )
    avg_time_to_first = req_with_first.aggregate(avg=Avg("time_to_first"))["avg"]
    avg_time_to_first_hours = None
    if avg_time_to_first:
        avg_time_to_first_hours = round(avg_time_to_first.total_seconds() / 3600, 2)

    # Pagination for approval logs
    from django.core.paginator import Paginator

    try:
        page_size = int(request.GET.get("page_size", 25))
    except (TypeError, ValueError):
        page_size = 25
    page_size = max(5, min(page_size, 200))
    page_number = request.GET.get("page", 1)
    paginator = Paginator(approval_trails, page_size)
    approval_logs_page = paginator.get_page(page_number)

    # Average approval time (time from creation to first approval)
    requisitions_with_approvals = Requisition.objects.filter(
        created_at__gte=start_date, status__in=["paid", "reviewed"]
    ).select_related("requested_by")

    context = {
        "approval_logs": approval_logs_page,
        "approver_stats": approver_stats,
        "action_breakdown": action_breakdown,
        "days": days,
        "total_logs": approval_trails.count(),
        "page_obj": approval_logs_page,
        "page_size": page_size,
        "avg_time_to_first_hours": avg_time_to_first_hours,
    }

    return render(request, "reports/approval_report.html", context)


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def approval_logs_export_csv(request):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    qs = (
        ApprovalTrail.objects.filter(timestamp__gte=start_date)
        .select_related("user", "requisition")
        .order_by("-timestamp")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="approval_logs_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(["Timestamp", "User", "Action", "Requisition", "Comment"])
    for a in qs.iterator():
        writer.writerow(
            [
                a.timestamp.strftime("%Y-%m-%d %H:%M"),
                getattr(a.user, "username", ""),
                a.action,
                a.requisition_id,
                (a.comment or "").replace("\n", " ").strip(),
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def approval_logs_export_xlsx(request):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    qs = (
        ApprovalTrail.objects.filter(timestamp__gte=start_date)
        .select_related("user", "requisition")
        .order_by("-timestamp")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Approval Logs"
    ws.append(["Timestamp", "User", "Action", "Requisition", "Comment"])
    for a in qs.iterator():
        ws.append(
            [
                a.timestamp.strftime("%Y-%m-%d %H:%M"),
                getattr(a.user, "username", ""),
                a.action,
                a.requisition_id,
                (a.comment or "").replace("\n", " ").strip(),
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="approval_logs_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def approver_perf_export_csv(request):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    approval_trails = ApprovalTrail.objects.filter(timestamp__gte=start_date)
    approver_stats = (
        approval_trails.values("user__username", "user__id")
        .annotate(
            total_actions=Count("id"),
            approvals=Count("id", filter=Q(action="approved")),
            rejections=Count("id", filter=Q(action="rejected")),
            change_requests=Count("id", filter=Q(action="changes_requested")),
        )
        .order_by("-total_actions")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="approver_performance_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(
        [
            "User",
            "User ID",
            "Total Actions",
            "Approvals",
            "Rejections",
            "Change Requests",
            "Approval Rate %",
        ]
    )
    for s in approver_stats:
        rate = (
            round((s["approvals"] / s["total_actions"] * 100.0), 2)
            if s["total_actions"]
            else 0
        )
        writer.writerow(
            [
                s["user__username"],
                s["user__id"],
                s["total_actions"],
                s["approvals"],
                s["rejections"],
                s["change_requests"],
                rate,
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def approver_perf_export_xlsx(request):
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    approval_trails = ApprovalTrail.objects.filter(timestamp__gte=start_date)
    approver_stats = (
        approval_trails.values("user__username", "user__id")
        .annotate(
            total_actions=Count("id"),
            approvals=Count("id", filter=Q(action="approved")),
            rejections=Count("id", filter=Q(action="rejected")),
            change_requests=Count("id", filter=Q(action="changes_requested")),
        )
        .order_by("-total_actions")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Approver Performance"
    ws.append(
        [
            "User",
            "User ID",
            "Total Actions",
            "Approvals",
            "Rejections",
            "Change Requests",
            "Approval Rate %",
        ]
    )
    for s in approver_stats:
        rate = (
            round((s["approvals"] / s["total_actions"] * 100.0), 2)
            if s["total_actions"]
            else 0
        )
        ws.append(
            [
                s["user__username"],
                s["user__id"],
                s["total_actions"],
                s["approvals"],
                s["rejections"],
                s["change_requests"],
                rate,
            ]
        )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="approver_performance_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_user_activity_report", app_label="reports")
def user_activity_report(request):
    """
    User activity report: requisition creation, approval actions, payment execution.
    """
    # Get date range
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)

    # User requisition activity
    user_requisition_stats = (
        User.objects.filter(requisition__created_at__gte=start_date)
        .annotate(
            total_requisitions=Count("requisition"),
            total_amount_requested=Sum("requisition__amount"),
            pending_count=Count(
                "requisition",
                filter=Q(
                    requisition__status__in=[
                        "pending",
                        "pending_dept_approval",
                        "pending_branch_approval",
                        "pending_regional_review",
                        "pending_finance_review",
                        "pending_treasury_validation",
                        "pending_cfo_approval",
                    ]
                ),
            ),
            paid_count=Count("requisition", filter=Q(requisition__status="paid")),
            rejected_count=Count(
                "requisition", filter=Q(requisition__status="rejected")
            ),
        )
        .filter(total_requisitions__gt=0)
        .order_by("-total_requisitions")
    )

    # User approval activity
    user_approval_stats = (
        User.objects.filter(approvaltrail__timestamp__gte=start_date)
        .annotate(
            total_approvals=Count(
                "approvaltrail", filter=Q(approvaltrail__action="approved")
            ),
            total_rejections=Count(
                "approvaltrail", filter=Q(approvaltrail__action="rejected")
            ),
            total_change_requests=Count(
                "approvaltrail", filter=Q(approvaltrail__action="changes_requested")
            ),
        )
        .filter(
            Q(total_approvals__gt=0)
            | Q(total_rejections__gt=0)
            | Q(total_change_requests__gt=0)
        )
        .order_by("-total_approvals")
    )

    # User payment execution activity
    user_payment_stats_qs = (
        User.objects.filter(executed_payments__requisition__created_at__gte=start_date)
        .annotate(
            total_payments=Count("executed_payments"),
            successful_payments=Count(
                "executed_payments", filter=Q(executed_payments__status="success")
            ),
            failed_payments=Count(
                "executed_payments", filter=Q(executed_payments__status="failed")
            ),
            total_amount_paid=Sum(
                "executed_payments__amount",
                filter=Q(executed_payments__status="success"),
            ),
        )
        .filter(total_payments__gt=0)
        .order_by("-total_payments")
    )

    # Pagination
    page_size = max(5, min(200, int(request.GET.get("page_size", 25))))

    req_paginator = Paginator(user_requisition_stats, page_size)
    req_page = req_paginator.get_page(request.GET.get("req_page", 1))

    appr_paginator = Paginator(user_approval_stats, page_size)
    appr_page = appr_paginator.get_page(request.GET.get("appr_page", 1))

    pay_paginator = Paginator(user_payment_stats_qs, page_size)
    pay_page = pay_paginator.get_page(request.GET.get("pay_page", 1))

    context = {
        "user_requisition_stats": req_page,
        "user_approval_stats": appr_page,
        "user_payment_stats": pay_page,
        "page_size": page_size,
        "days": days,
    }

    return render(request, "reports/user_activity_report.html", context)


@require_app_access("reports")
@require_permission("view_budget_vs_actuals_report", app_label="reports")
def budget_vs_actuals_report(request):
    """Budget vs Actuals by Cost Center (monthly), with variance."""
    from organization.models import Branch, CostCenter, Department
    from reports.models import BudgetAllocation

    try:
        year = int(request.GET.get("year", timezone.now().year))
    except (TypeError, ValueError):
        year = timezone.now().year
    group = request.GET.get("group", "cost_center")  # cost_center | department | branch

    # Actuals (sum of amounts) in the given year by month and scope
    reqs = Requisition.objects.filter(created_at__year=year).select_related(
        "cost_center", "department", "branch"
    )

    if group == "department":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "department_id", "department__name")
            .annotate(actual=Sum("amount"))
            .order_by("month", "department__name")
        )
    elif group == "branch":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "branch_id", "branch__name")
            .annotate(actual=Sum("amount"))
            .order_by("month", "branch__name")
        )
    else:
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values(
                "month",
                "cost_center_id",
                "cost_center__name",
                "department__name",
                "branch__name",
            )
            .annotate(actual=Sum("amount"))
            .order_by("month", "cost_center__name")
        )

    # Collect keys for budget lookup
    months = set()
    rows = []
    for a in actuals:
        month_num = a["month"].month if a["month"] else None
        months.add(month_num)
        rows.append(a)

    # Fetch budgets for the same scope
    budgets_qs = BudgetAllocation.objects.filter(year=year)
    budget_map = {}
    for b in budgets_qs.iterator():
        key = (
            ("branch", b.branch_id)
            if group == "branch"
            else (
                ("department", b.department_id)
                if group == "department"
                else ("cost_center", b.cost_center_id)
            )
        )
        month_key = b.month or 0
        budget_map.setdefault((key[0], key[1], month_key), 0)
        budget_map[(key[0], key[1], month_key)] += float(b.amount)

    # Build result with variance
    data = []
    for r in rows:
        if group == "branch":
            scope_key = ("branch", r.get("branch_id"))
            scope_name = r.get("branch__name") or "-"
        elif group == "department":
            scope_key = ("department", r.get("department_id"))
            scope_name = r.get("department__name") or "-"
        else:
            scope_key = ("cost_center", r.get("cost_center_id"))
            scope_name = r.get("cost_center__name") or "-"

        month_num = r["month"].month if r["month"] else 0
        budget = budget_map.get(
            (scope_key[0], scope_key[1], month_num),
            budget_map.get((scope_key[0], scope_key[1], 0), 0),
        )
        actual = float(r["actual"] or 0)
        variance = actual - budget
        variance_pct = (variance / budget * 100.0) if budget else None
        data.append(
            {
                "scope_name": scope_name,
                "month": r["month"],
                "actual": actual,
                "budget": budget,
                "variance": variance,
                "variance_pct": variance_pct,
            }
        )

    # Sort by scope then month
    data.sort(key=lambda x: (x["scope_name"] or "", x["month"]))

    # Pagination
    page_size = max(5, min(200, int(request.GET.get("page_size", 25))))
    paginator = Paginator(data, page_size)
    page_number = request.GET.get("page", 1)
    rows = paginator.get_page(page_number)

    context = {
        "year": year,
        "group": group,
        "rows": rows,
        "page_size": page_size,
    }
    return render(request, "reports/budget_vs_actuals.html", context)


@require_app_access("reports")
@require_permission("view_budget_vs_actuals_report", app_label="reports")
def budget_vs_actuals_export_csv(request):
    year = int(request.GET.get("year", timezone.now().year))
    request.GET._mutable = True  # safe in view scope
    request.GET["group"] = request.GET.get("group", "cost_center")
    response = budget_vs_actuals_report(request)
    # Re-run logic more directly for CSV to avoid template
    group = request.GET["group"]
    from reports.models import BudgetAllocation

    reqs = Requisition.objects.filter(created_at__year=year).select_related(
        "cost_center", "department", "branch"
    )
    if group == "department":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "department_id", "department__name")
            .annotate(actual=Sum("amount"))
        )
    elif group == "branch":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "branch_id", "branch__name")
            .annotate(actual=Sum("amount"))
        )
    else:
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "cost_center_id", "cost_center__name")
            .annotate(actual=Sum("amount"))
        )
    budgets_qs = BudgetAllocation.objects.filter(year=year)
    budget_map = {}
    for b in budgets_qs.iterator():
        key = (
            ("branch", b.branch_id)
            if group == "branch"
            else (
                ("department", b.department_id)
                if group == "department"
                else ("cost_center", b.cost_center_id)
            )
        )
        month_key = b.month or 0
        budget_map.setdefault((key[0], key[1], month_key), 0)
        budget_map[(key[0], key[1], month_key)] += float(b.amount)
    # Build CSV
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="budget_vs_actuals_{year}.csv"'
    writer = csv.writer(resp)
    writer.writerow(["Scope", "Month", "Actual", "Budget", "Variance", "Variance %"])
    for r in actuals:
        scope_name = (
            r.get("branch__name")
            or r.get("department__name")
            or r.get("cost_center__name")
            or "-"
        )
        month_num = r["month"].month if r["month"] else 0
        budget = (
            budget_map.get(("branch", r.get("branch_id"), month_num), 0)
            if group == "branch"
            else (
                budget_map.get(("department", r.get("department_id"), month_num), 0)
                if group == "department"
                else budget_map.get(
                    ("cost_center", r.get("cost_center_id"), month_num), 0
                )
            )
        )
        actual = float(r["actual"] or 0)
        variance = actual - budget
        variance_pct = round((variance / budget * 100.0), 2) if budget else None
        writer.writerow(
            [
                scope_name,
                month_num,
                actual,
                budget,
                variance,
                variance_pct if variance_pct is not None else "",
            ]
        )
    return resp


@require_app_access("reports")
@require_permission("view_budget_vs_actuals_report", app_label="reports")
def budget_vs_actuals_export_xlsx(request):
    year = int(request.GET.get("year", timezone.now().year))
    group = request.GET.get("group", "cost_center")
    from reports.models import BudgetAllocation

    reqs = Requisition.objects.filter(created_at__year=year).select_related(
        "cost_center", "department", "branch"
    )
    if group == "department":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "department_id", "department__name")
            .annotate(actual=Sum("amount"))
        )
    elif group == "branch":
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "branch_id", "branch__name")
            .annotate(actual=Sum("amount"))
        )
    else:
        actuals = (
            reqs.annotate(month=TruncMonth("created_at"))
            .values("month", "cost_center_id", "cost_center__name")
            .annotate(actual=Sum("amount"))
        )
    budgets_qs = BudgetAllocation.objects.filter(year=year)
    budget_map = {}
    for b in budgets_qs.iterator():
        key = (
            ("branch", b.branch_id)
            if group == "branch"
            else (
                ("department", b.department_id)
                if group == "department"
                else ("cost_center", b.cost_center_id)
            )
        )
        month_key = b.month or 0
        budget_map.setdefault((key[0], key[1], month_key), 0)
        budget_map[(key[0], key[1], month_key)] += float(b.amount)
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget vs Actuals"
    ws.append(["Scope", "Month", "Actual", "Budget", "Variance", "Variance %"])
    for r in actuals:
        scope_name = (
            r.get("branch__name")
            or r.get("department__name")
            or r.get("cost_center__name")
            or "-"
        )
        month_num = r["month"].month if r["month"] else 0
        budget = (
            budget_map.get(("branch", r.get("branch_id"), month_num), 0)
            if group == "branch"
            else (
                budget_map.get(("department", r.get("department_id"), month_num), 0)
                if group == "department"
                else budget_map.get(
                    ("cost_center", r.get("cost_center_id"), month_num), 0
                )
            )
        )
        actual = float(r["actual"] or 0)
        variance = actual - budget
        variance_pct = round((variance / budget * 100.0), 2) if budget else None
        ws.append(
            [
                scope_name,
                month_num,
                actual,
                budget,
                variance,
                variance_pct if variance_pct is not None else "",
            ]
        )
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = (
        f'attachment; filename="budget_vs_actuals_{year}.xlsx"'
    )
    wb.save(resp)
    return resp


@require_app_access("reports")
@require_permission("view_stuck_approvals_report", app_label="reports")
def stuck_approvals_report(request):
    """List requisitions stuck in pending statuses beyond a threshold number of days."""
    older_than = int(request.GET.get("older_than", 7))
    status_filter = request.GET.get(
        "status", "pending_only"
    )  # 'pending_only' or 'all_pending_*'
    start_cutoff = timezone.now() - timedelta(days=older_than)

    qs = Requisition.objects.filter(created_at__lt=start_cutoff).select_related(
        "requested_by", "branch", "department"
    )
    if status_filter == "pending_only":
        qs = qs.filter(status__startswith="pending")
    else:
        qs = qs.filter(
            status__in=[
                s for s, _ in Requisition.STATUS_CHOICES if s.startswith("pending")
            ]
        )

    qs = qs.order_by("created_at")

    # Aggregates by status
    by_status = (
        qs.values("status").annotate(count=Count("transaction_id")).order_by("-count")
    )

    # Helper to compute age in days
    def age_days(dt):
        return (timezone.now() - dt).days

    items_data = [
        {
            "transaction_id": r.transaction_id,
            "requested_by": getattr(r.requested_by, "username", ""),
            "branch": getattr(r.branch, "name", None),
            "department": getattr(r.department, "name", None),
            "amount": r.amount,
            "status": r.status,
            "created_at": r.created_at,
            "age_days": age_days(r.created_at),
            "next_approver": getattr(r.next_approver, "username", None),
        }
        for r in qs
    ]

    # Pagination
    page_size = max(5, min(200, int(request.GET.get("page_size", 25))))
    paginator = Paginator(items_data, page_size)
    page_number = request.GET.get("page", 1)
    items = paginator.get_page(page_number)

    context = {
        "older_than": older_than,
        "status_filter": status_filter,
        "by_status": by_status,
        "items": items,
        "page_size": page_size,
        "total": len(items_data),
    }
    return render(request, "reports/stuck_approvals.html", context)


@require_app_access("reports")
@require_permission("view_threshold_overrides_report", app_label="reports")
def threshold_overrides_report(request):
    """Requisitions that exceeded or bypassed configured approval thresholds."""
    days = int(request.GET.get("days", 90))
    start_date = timezone.now() - timedelta(days=days)

    # Subquery: requisitions with any override/skipped roles in approval trail
    trail_overrides_qs = ApprovalTrail.objects.filter(
        requisition_id=OuterRef("pk")
    ).filter(Q(override=True) | Q(skipped_roles__isnull=False))

    # Build base filter for out-of-range amounts relative to applied threshold
    out_of_range_q = Q(applied_threshold__isnull=False) & (
        Q(amount__lt=F("applied_threshold__min_amount"))
        | Q(amount__gt=F("applied_threshold__max_amount"))
    )

    qs = (
        Requisition.objects.filter(created_at__gte=start_date)
        .select_related("requested_by", "branch", "department", "applied_threshold")
        .annotate(has_trail_override=Exists(trail_overrides_qs))
        .filter(
            out_of_range_q
            | Q(is_fast_tracked=True)
            | Q(
                pk__in=Subquery(
                    ApprovalTrail.objects.filter(
                        Q(override=True) | Q(skipped_roles__isnull=False)
                    ).values("requisition_id")
                )
            )
        )
        .order_by("-created_at")
    )

    # Categorize
    by_category = {
        "out_of_range": qs.filter(out_of_range_q).count(),
        "fast_tracked": qs.filter(is_fast_tracked=True).count(),
        "trail_override": qs.filter(has_trail_override=True).count(),
    }

    items_data = [
        {
            "transaction_id": r.transaction_id,
            "requested_by": getattr(r.requested_by, "username", ""),
            "branch": getattr(r.branch, "name", None),
            "department": getattr(r.department, "name", None),
            "amount": r.amount,
            "status": r.status,
            "created_at": r.created_at,
            "applied_threshold": getattr(r.applied_threshold, "name", None),
            "min_amount": getattr(r.applied_threshold, "min_amount", None),
            "max_amount": getattr(r.applied_threshold, "max_amount", None),
            "is_fast_tracked": r.is_fast_tracked,
            "has_trail_override": r.has_trail_override,
        }
        for r in qs
    ]

    # Pagination
    page_size = max(5, min(200, int(request.GET.get("page_size", 25))))
    paginator = Paginator(items_data, page_size)
    page_number = request.GET.get("page", 1)
    items = paginator.get_page(page_number)

    context = {
        "days": days,
        "items": items,
        "page_size": page_size,
        "total": len(items_data),
        "by_category": by_category,
    }
    return render(request, "reports/threshold_overrides.html", context)


@require_app_access("reports")
@require_permission("view_threshold_overrides_report", app_label="reports")
def threshold_overrides_export_csv(request):
    days = int(request.GET.get("days", 90))
    start_date = timezone.now() - timedelta(days=days)
    out_of_range_q = Q(applied_threshold__isnull=False) & (
        Q(amount__lt=F("applied_threshold__min_amount"))
        | Q(amount__gt=F("applied_threshold__max_amount"))
    )
    ovr_ids = ApprovalTrail.objects.filter(
        Q(override=True) | Q(skipped_roles__isnull=False)
    ).values("requisition_id")
    qs = (
        Requisition.objects.filter(created_at__gte=start_date)
        .select_related("requested_by", "branch", "department", "applied_threshold")
        .filter(out_of_range_q | Q(is_fast_tracked=True) | Q(pk__in=Subquery(ovr_ids)))
        .order_by("-created_at")
    )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="threshold_overrides_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(
        [
            "Transaction ID",
            "Requested By",
            "Branch",
            "Department",
            "Amount",
            "Status",
            "Created",
            "Applied Threshold",
            "Min",
            "Max",
            "Fast Tracked",
            "Trail Override",
        ]
    )
    for r in qs.iterator():
        writer.writerow(
            [
                r.transaction_id,
                getattr(r.requested_by, "username", ""),
                getattr(r.branch, "name", "") if r.branch else "",
                getattr(r.department, "name", "") if r.department else "",
                f"{r.amount}",
                r.status,
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                getattr(r.applied_threshold, "name", ""),
                getattr(r.applied_threshold, "min_amount", ""),
                getattr(r.applied_threshold, "max_amount", ""),
                "Yes" if r.is_fast_tracked else "No",
                (
                    "Yes"
                    if ApprovalTrail.objects.filter(requisition_id=r.pk)
                    .filter(Q(override=True) | Q(skipped_roles__isnull=False))
                    .exists()
                    else "No"
                ),
            ]
        )
    return response


# ================================================================================
# PHASE 1: QUICK WIN REPORTS
# ================================================================================

EXPENSE_CATEGORIES = {
    "travel": "Travel & Transport",
    "supplies": "Office Supplies",
    "services": "Professional Services",
    "utilities": "Utilities",
    "maintenance": "Maintenance & Repairs",
    "communication": "Communication",
    "marketing": "Marketing & Advertising",
    "training": "Training & Development",
    "entertainment": "Entertainment & Hospitality",
    "equipment": "Equipment & Assets",
    "other": "Other Expenses",
}


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def category_spending_report(request):
    """Category spending analysis with breakdown by department/branch."""
    days = int(request.GET.get("days", 30))
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    start_date = timezone.now() - timedelta(days=days)

    qs = Requisition.objects.filter(created_at__gte=start_date).exclude(status="draft")
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    if department_id:
        qs = qs.filter(department_id=department_id)

    # Overall stats
    total_count = qs.count()
    total_amount = qs.aggregate(Sum("amount"))["amount__sum"] or 0
    avg_amount = qs.aggregate(Avg("amount"))["amount__avg"] or 0

    # Pagination size
    try:
        page_size = int(request.GET.get("page_size", 25))
    except (TypeError, ValueError):
        page_size = 25
    page_size = max(5, min(page_size, 200))

    # Category breakdown (using purpose as spending category)
    category_stats_qs = (
        qs.values("purpose")
        .annotate(
            count=Count("transaction_id"), total=Sum("amount"), average=Avg("amount")
        )
        .order_by("-total")
    )
    # Pagination for category stats
    from django.core.paginator import Paginator

    page = request.GET.get("page", 1)
    cat_paginator = Paginator(list(category_stats_qs), page_size)
    category_stats = cat_paginator.get_page(page)

    # Add display names and percentages
    for cat in category_stats:
        cat["display_name"] = cat.get("purpose") or "Unspecified"
        cat["percentage"] = (cat["total"] / total_amount * 100) if total_amount else 0

    top_category = category_stats[0] if category_stats else {}
    if top_category:
        top_category["name"] = top_category.get("display_name")

    # Department breakdown
    by_department_qs = (
        qs.values("department__name", "purpose")
        .annotate(count=Count("transaction_id"), total=Sum("amount"))
        .order_by("department__name", "-total")
    )
    dept_page = request.GET.get("dept_page", 1)
    dept_paginator = Paginator(list(by_department_qs), page_size)
    by_department = dept_paginator.get_page(dept_page)

    for item in by_department:
        item["display_name"] = item.get("purpose") or "Unspecified"

    from organization.models import Branch, Department

    branches = Branch.objects.all().order_by("name")
    departments = Department.objects.all().order_by("name")

    context = {
        "days": days,
        "total_count": total_count,
        "total_amount": total_amount,
        "avg_amount": avg_amount,
        "category_stats": category_stats,
        "cat_paginator": cat_paginator,
        "dept_paginator": dept_paginator,
        "top_category": top_category,
        "by_department": by_department,
        "branches": branches,
        "departments": departments,
        "selected_branch": branch_id,
        "selected_department": department_id,
    }
    return render(request, "reports/category_spending.html", context)


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def category_spending_export_csv(request):
    """Export category spending to CSV."""
    days = int(request.GET.get("days", 30))
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    start_date = timezone.now() - timedelta(days=days)

    qs = Requisition.objects.filter(created_at__gte=start_date).exclude(status="draft")
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    if department_id:
        qs = qs.filter(department_id=department_id)

    category_stats = (
        qs.values("purpose")
        .annotate(
            count=Count("transaction_id"), total=Sum("amount"), average=Avg("amount")
        )
        .order_by("-total")
    )

    total_amount = qs.aggregate(Sum("amount"))["amount__sum"] or 0

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="category_spending_{timezone.now().date()}.csv"'
    )
    writer = csv.writer(response)
    writer.writerow(
        ["Category", "Count", "Total Amount", "Average Amount", "Percentage"]
    )
    for cat in category_stats:
        display_name = cat.get("purpose") or "Unspecified"
        percentage = (cat["total"] / total_amount * 100) if total_amount else 0
        writer.writerow(
            [
                display_name,
                cat["count"],
                f"{cat['total']}",
                f"{cat['average']}",
                f"{percentage:.1f}%",
            ]
        )
    return response


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def category_spending_export_xlsx(request):
    """Export category spending to Excel."""
    days = int(request.GET.get("days", 30))
    branch_id = request.GET.get("branch", "")
    department_id = request.GET.get("department", "")
    start_date = timezone.now() - timedelta(days=days)

    qs = Requisition.objects.filter(created_at__gte=start_date).exclude(status="draft")
    if branch_id:
        qs = qs.filter(branch_id=branch_id)
    if department_id:
        qs = qs.filter(department_id=department_id)

    category_stats = (
        qs.values("purpose")
        .annotate(
            count=Count("transaction_id"), total=Sum("amount"), average=Avg("amount")
        )
        .order_by("-total")
    )

    total_amount = qs.aggregate(Sum("amount"))["amount__sum"] or 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Category Spending"
    ws.append(["Category", "Count", "Total Amount", "Average Amount", "Percentage"])
    for cat in category_stats:
        display_name = cat.get("purpose") or "Unspecified"
        percentage = (cat["total"] / total_amount * 100) if total_amount else 0
        ws.append(
            [
                display_name,
                cat["count"],
                float(cat["total"]),
                float(cat["average"]),
                f"{percentage:.1f}%",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="category_spending_{timezone.now().date()}.xlsx"'
    )
    wb.save(response)
    return response


@require_app_access("reports")
@require_permission("view_treasury_report", app_label="reports")
def payment_method_analysis(request):
    """Payment method performance analysis."""
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    page_size = int(request.GET.get("page_size", 25))
    page_size = max(5, min(page_size, 200))

    payments = Payment.objects.filter(created_at__gte=start_date).select_related(
        "requisition"
    )

    # Overall stats
    total_payments = payments.count()
    total_amount = payments.aggregate(Sum("amount"))["amount__sum"] or 0
    success_count = payments.filter(status="success").count()
    failed_count = payments.filter(status="failed").count()
    success_rate = (success_count / total_payments * 100) if total_payments else 0

    # By method
    by_method = (
        payments.values("method")
        .annotate(
            count=Count("payment_id"),
            total_amount=Sum("amount"),
            success_count=Count("payment_id", filter=Q(status="success")),
            failed_count=Count("payment_id", filter=Q(status="failed")),
            avg_amount=Avg("amount"),
        )
        .order_by("-total_amount")
    )
    # Pagination for methods
    from django.core.paginator import Paginator

    method_page = request.GET.get("page", 1)
    method_paginator = Paginator(list(by_method), page_size)
    by_method = method_paginator.get_page(method_page)

    for item in by_method:
        item["success_rate"] = (
            (item["success_count"] / item["count"] * 100) if item["count"] else 0
        )
        item["method_display"] = dict(Payment.PAYMENT_METHOD_CHOICES).get(
            item["method"], item["method"]
        )

    # Processing time (successful payments only)
    from django.db.models import DurationField, ExpressionWrapper, F

    payments_with_time = Payment.objects.filter(
        created_at__gte=start_date,
        status="success",
        requisition__created_at__isnull=False,
    ).annotate(
        processing_time=ExpressionWrapper(
            F("created_at") - F("requisition__created_at"), output_field=DurationField()
        )
    )
    avg_processing = payments_with_time.aggregate(avg=Avg("processing_time"))["avg"]
    avg_processing_hours = (
        avg_processing.total_seconds() / 3600 if avg_processing else 0
    )

    # Failure reasons (if available)
    failed_payments_qs = (
        payments.filter(status="failed")
        .values("method")
        .annotate(count=Count("payment_id"))
    )
    failed_page = request.GET.get("failed_page", 1)
    failed_paginator = Paginator(list(failed_payments_qs), page_size)
    failed_payments = failed_paginator.get_page(failed_page)

    context = {
        "days": days,
        "page_size": page_size,
        "total_payments": total_payments,
        "total_amount": total_amount,
        "success_count": success_count,
        "failed_count": failed_count,
        "success_rate": success_rate,
        "by_method": by_method,
        "method_paginator": method_paginator,
        "avg_processing_hours": avg_processing_hours,
        "failed_payments": failed_payments,
        "failed_paginator": failed_paginator,
    }
    return render(request, "reports/payment_method_analysis.html", context)


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def regional_comparison_report(request):
    """Compare performance across regions/branches."""
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    page_size = int(request.GET.get("page_size", 25))
    page_size = max(5, min(page_size, 200))

    # Branch comparison
    by_branch_qs = (
        Requisition.objects.filter(created_at__gte=start_date)
        .exclude(status="draft")
        .values("branch__name", "branch__id")
        .annotate(
            count=Count("transaction_id"),
            total_amount=Sum("amount"),
            avg_amount=Avg("amount"),
            paid_count=Count("transaction_id", filter=Q(status="paid")),
            pending_count=Count(
                "transaction_id", filter=Q(status__startswith="pending")
            ),
            rejected_count=Count("transaction_id", filter=Q(status="rejected")),
        )
        .order_by("-total_amount")
    )
    from django.core.paginator import Paginator

    branch_page = request.GET.get("page", 1)
    branch_paginator = Paginator(list(by_branch_qs), page_size)
    by_branch = branch_paginator.get_page(branch_page)

    for item in by_branch:
        item["completion_rate"] = (
            (item["paid_count"] / item["count"] * 100) if item["count"] else 0
        )

    # Payment success by branch
    branch_payments_qs = (
        Payment.objects.filter(created_at__gte=start_date)
        .values("requisition__branch__name")
        .annotate(
            total_payments=Count("payment_id"),
            successful=Count("payment_id", filter=Q(status="success")),
            failed=Count("payment_id", filter=Q(status="failed")),
        )
        .order_by("-total_payments")
    )
    pay_page = request.GET.get("pay_page", 1)
    pay_paginator = Paginator(list(branch_payments_qs), page_size)
    branch_payments = pay_paginator.get_page(pay_page)

    for item in branch_payments:
        item["success_rate"] = (
            (item["successful"] / item["total_payments"] * 100)
            if item["total_payments"]
            else 0
        )

    # Approval time by branch
    from django.db.models import DurationField, ExpressionWrapper, F, Min

    branch_approval_time_qs = (
        Requisition.objects.filter(created_at__gte=start_date, status="paid")
        .values("branch__name")
        .annotate(first_approval=Min("approvaltrail__timestamp"))
        .annotate(
            avg_time=Avg(
                ExpressionWrapper(
                    F("approvaltrail__timestamp") - F("created_at"),
                    output_field=DurationField(),
                )
            )
        )
    )
    appr_page = request.GET.get("appr_page", 1)
    appr_paginator = Paginator(list(branch_approval_time_qs), page_size)
    branch_approval_time = appr_paginator.get_page(appr_page)

    for item in branch_approval_time:
        if item["avg_time"]:
            item["avg_hours"] = item["avg_time"].total_seconds() / 3600

    context = {
        "days": days,
        "page_size": page_size,
        "by_branch": by_branch,
        "branch_payments": branch_payments,
        "branch_approval_time": branch_approval_time,
        "branch_paginator": branch_paginator,
        "pay_paginator": pay_paginator,
        "appr_paginator": appr_paginator,
    }
    return render(request, "reports/regional_comparison.html", context)


@require_app_access("reports")
@require_permission("view_approval_report", app_label="reports")
def rejection_analysis_report(request):
    """Analyze rejection patterns and reasons."""
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    page_size = int(request.GET.get("page_size", 25))
    page_size = max(5, min(page_size, 200))

    # Rejected requisitions
    rejected = Requisition.objects.filter(
        created_at__gte=start_date, status="rejected"
    ).select_related("requested_by", "branch", "department")

    total_rejected = rejected.count()
    total_requisitions = (
        Requisition.objects.filter(created_at__gte=start_date)
        .exclude(status="draft")
        .count()
    )
    rejection_rate = (
        (total_rejected / total_requisitions * 100) if total_requisitions else 0
    )

    # Rejection trails
    rejection_trails = ApprovalTrail.objects.filter(
        requisition__created_at__gte=start_date, action="rejected"
    ).select_related("user", "requisition")

    # By category (using purpose field available on Requisition)
    by_category_qs = (
        rejected.values("purpose")
        .annotate(count=Count("transaction_id"), total_amount=Sum("amount"))
        .order_by("-count")
    )
    from django.core.paginator import Paginator

    page = request.GET.get("page", 1)
    cat_paginator = Paginator(list(by_category_qs), page_size)
    by_category = cat_paginator.get_page(page)

    for cat in by_category:
        cat["display_name"] = cat.get("purpose") or "Unspecified"

    # By department
    by_department_qs = (
        rejected.values("department__name")
        .annotate(count=Count("transaction_id"))
        .order_by("-count")
    )
    dept_page = request.GET.get("dept_page", 1)
    dept_paginator = Paginator(list(by_department_qs), page_size)
    by_department = dept_paginator.get_page(dept_page)

    # By rejector
    by_rejector_qs = (
        rejection_trails.values("user__username", "user__role")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    rej_page = request.GET.get("rej_page", 1)
    rej_paginator = Paginator(list(by_rejector_qs), page_size)
    by_rejector = rej_paginator.get_page(rej_page)

    # Top rejection reasons (from comments)
    rejection_reasons = (
        rejection_trails.exclude(comment__isnull=True)
        .exclude(comment="")
        .values("comment")[:10]
    )

    context = {
        "days": days,
        "page_size": page_size,
        "total_rejected": total_rejected,
        "rejection_rate": rejection_rate,
        "by_category": by_category,
        "by_department": by_department,
        "by_rejector": by_rejector,
        "rejection_reasons": rejection_reasons,
        "cat_paginator": cat_paginator,
        "dept_paginator": dept_paginator,
        "rej_paginator": rej_paginator,
    }
    return render(request, "reports/rejection_analysis.html", context)


@require_app_access("reports")
@require_permission("view_transaction_report", app_label="reports")
def average_metrics_report(request):
    """Average requisition metrics across different dimensions."""
    days = int(request.GET.get("days", 30))
    start_date = timezone.now() - timedelta(days=days)
    page_size = int(request.GET.get("page_size", 25))
    page_size = max(5, min(page_size, 200))

    qs = Requisition.objects.filter(created_at__gte=start_date).exclude(status="draft")

    # Overall averages
    overall = qs.aggregate(
        avg_amount=Avg("amount"),
        total_count=Count("transaction_id"),
        total_amount=Sum("amount"),
    )

    # By department
    by_department_qs = (
        qs.values("department__name")
        .annotate(
            avg_amount=Avg("amount"), count=Count("transaction_id"), total=Sum("amount")
        )
        .order_by("-total")
    )
    from django.core.paginator import Paginator

    dept_page = request.GET.get("dept_page", 1)
    dept_paginator = Paginator(list(by_department_qs), page_size)
    by_department = dept_paginator.get_page(dept_page)

    # By branch
    by_branch_qs = (
        qs.values("branch__name")
        .annotate(
            avg_amount=Avg("amount"), count=Count("transaction_id"), total=Sum("amount")
        )
        .order_by("-total")
    )
    branch_page = request.GET.get("branch_page", 1)
    branch_paginator = Paginator(list(by_branch_qs), page_size)
    by_branch = branch_paginator.get_page(branch_page)

    # By user (top requesters)
    by_user_qs = (
        qs.values("requested_by__username", "requested_by__department__name")
        .annotate(
            avg_amount=Avg("amount"), count=Count("transaction_id"), total=Sum("amount")
        )
        .order_by("-count")
    )
    user_page = request.GET.get("user_page", 1)
    user_paginator = Paginator(list(by_user_qs), page_size)
    by_user = user_paginator.get_page(user_page)

    # Approval time metrics
    from django.db.models import DurationField, ExpressionWrapper, F, Min

    paid_reqs = qs.filter(status="paid")

    first_approval_subq = (
        ApprovalTrail.objects.filter(requisition_id=OuterRef("pk"), action="approved")
        .order_by("timestamp")
        .values("timestamp")[:1]
    )

    reqs_with_approval = (
        paid_reqs.annotate(first_approval_at=Subquery(first_approval_subq))
        .filter(first_approval_at__isnull=False)
        .annotate(
            approval_time=ExpressionWrapper(
                F("first_approval_at") - F("created_at"), output_field=DurationField()
            )
        )
    )

    avg_approval_time = reqs_with_approval.aggregate(avg=Avg("approval_time"))["avg"]
    avg_approval_hours = (
        avg_approval_time.total_seconds() / 3600 if avg_approval_time else 0
    )

    # Payment time (from creation to payment)
    payment_subq = (
        Payment.objects.filter(requisition_id=OuterRef("pk"), status="success")
        .order_by("created_at")
        .values("created_at")[:1]
    )

    reqs_with_payment = (
        paid_reqs.annotate(payment_at=Subquery(payment_subq))
        .filter(payment_at__isnull=False)
        .annotate(
            payment_time=ExpressionWrapper(
                F("payment_at") - F("created_at"), output_field=DurationField()
            )
        )
    )

    avg_payment_time = reqs_with_payment.aggregate(avg=Avg("payment_time"))["avg"]
    avg_payment_hours = (
        avg_payment_time.total_seconds() / 3600 if avg_payment_time else 0
    )

    context = {
        "days": days,
        "page_size": page_size,
        "overall": overall,
        "by_department": by_department,
        "by_branch": by_branch,
        "by_user": by_user,
        "avg_approval_hours": avg_approval_hours,
        "avg_payment_hours": avg_payment_hours,
        "dept_paginator": dept_paginator,
        "branch_paginator": branch_paginator,
        "user_paginator": user_paginator,
    }
    return render(request, "reports/average_metrics.html", context)
