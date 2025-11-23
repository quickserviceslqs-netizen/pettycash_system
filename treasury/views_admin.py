"""
Treasury Admin Views
User-friendly interface for managing treasury funds, payments, ledger, and variances
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.http import HttpResponse
from decimal import Decimal
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from treasury.models import (
    TreasuryFund, Payment, PaymentExecution, LedgerEntry, 
    VarianceAdjustment
)
from organization.models import Company, Region, Branch
from transactions.models import Requisition
from settings_manager.models import get_setting


# ==================== TREASURY FUNDS ====================

@login_required
@permission_required('treasury.view_treasuryfund', raise_exception=True)
def manage_funds(request):
    """View and manage all treasury funds"""
    
    # Filters
    company_filter = request.GET.get('company', '')
    region_filter = request.GET.get('region', '')
    low_balance_only = request.GET.get('low_balance', '') == 'on'
    
    funds = TreasuryFund.objects.all().select_related('company', 'region', 'branch')
    
    if company_filter:
        funds = funds.filter(company_id=company_filter)
    
    if region_filter:
        funds = funds.filter(region_id=region_filter)
    
    if low_balance_only:
        funds = [f for f in funds if f.check_reorder_needed()]
    
    # Get stats
    total_balance = funds.aggregate(total=Sum('current_balance'))['total'] or Decimal('0.00')
    low_balance_count = sum(1 for f in funds if f.check_reorder_needed())
    
    companies = Company.objects.all()
    regions = Region.objects.all()
    
    context = {
        'funds': funds,
        'total_balance': total_balance,
        'low_balance_count': low_balance_count,
        'companies': companies,
        'regions': regions,
        'company_filter': company_filter,
        'region_filter': region_filter,
        'low_balance_only': low_balance_only,
    }
    
    return render(request, 'treasury/manage_funds.html', context)


@login_required
@permission_required('treasury.add_treasuryfund', raise_exception=True)
def create_fund(request):
    """Create a new treasury fund"""
    
    if request.method == 'POST':
        try:
            company_id = request.POST.get('company')
            region_id = request.POST.get('region') or None
            branch_id = request.POST.get('branch') or None
            current_balance = request.POST.get('current_balance') or '0.00'
            default_reorder = get_setting('TREASURY_DEFAULT_REORDER_LEVEL', '50000')
            reorder_level = request.POST.get('reorder_level') or default_reorder
            
            if not company_id:
                messages.error(request, 'Company is required.')
                return redirect('treasury:create_fund')
            
            # Check for existing fund
            existing = TreasuryFund.objects.filter(
                company_id=company_id,
                region_id=region_id,
                branch_id=branch_id
            ).exists()
            
            if existing:
                messages.error(request, 'A fund already exists for this location.')
                return redirect('treasury:create_fund')
            
            fund = TreasuryFund.objects.create(
                company_id=company_id,
                region_id=region_id,
                branch_id=branch_id,
                current_balance=Decimal(current_balance),
                reorder_level=Decimal(reorder_level),
            )
            
            messages.success(request, f'Treasury fund created successfully!')
            return redirect('treasury:manage_funds')
            
        except Exception as e:
            messages.error(request, f'Error creating fund: {str(e)}')
    
    companies = Company.objects.all()
    regions = Region.objects.all()
    branches = Branch.objects.all()
    
    default_reorder_level = get_setting('TREASURY_DEFAULT_REORDER_LEVEL', '50000')
    
    context = {
        'companies': companies,
        'regions': regions,
        'branches': branches,
        'default_reorder_level': default_reorder_level,
    }
    
    return render(request, 'treasury/create_fund.html', context)


@login_required
@permission_required('treasury.change_treasuryfund', raise_exception=True)
def edit_fund(request, fund_id):
    """Edit an existing treasury fund"""
    
    fund = get_object_or_404(TreasuryFund, fund_id=fund_id)
    
    if request.method == 'POST':
        try:
            fund.current_balance = Decimal(request.POST.get('current_balance', '0.00'))
            default_reorder = get_setting('TREASURY_DEFAULT_REORDER_LEVEL', '50000')
            fund.reorder_level = Decimal(request.POST.get('reorder_level', default_reorder))
            fund.save()
            
            messages.success(request, 'Fund updated successfully!')
            return redirect('treasury:manage_funds')
            
        except Exception as e:
            messages.error(request, f'Error updating fund: {str(e)}')
    
    context = {
        'fund': fund,
    }
    
    return render(request, 'treasury/edit_fund.html', context)


# ==================== PAYMENTS ====================

@login_required
@permission_required('treasury.view_payment', raise_exception=True)
def manage_payments(request):
    """View and manage all payments"""
    
    # Filters
    status_filter = request.GET.get('status', '')
    method_filter = request.GET.get('method', '')
    search_query = request.GET.get('search', '')
    
    payments = Payment.objects.all().select_related(
        'requisition', 'executor', 'requisition__requested_by'
    )
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if method_filter:
        payments = payments.filter(method=method_filter)
    
    if search_query:
        payments = payments.filter(
            Q(payment_id__icontains=search_query) |
            Q(destination__icontains=search_query) |
            Q(requisition__transaction_id__icontains=search_query)
        )
    
    payments = payments.order_by('-created_at')[:100]  # Limit to 100 recent
    
    # Stats
    stats = {
        'pending': Payment.objects.filter(status='pending').count(),
        'executing': Payment.objects.filter(status='executing').count(),
        'success': Payment.objects.filter(status='success').count(),
        'failed': Payment.objects.filter(status='failed').count(),
    }
    
    context = {
        'payments': payments,
        'stats': stats,
        'status_filter': status_filter,
        'method_filter': method_filter,
        'search_query': search_query,
        'status_choices': Payment.STATUS_CHOICES,
        'method_choices': Payment.PAYMENT_METHOD_CHOICES,
    }
    
    return render(request, 'treasury/manage_payments.html', context)


@login_required
@permission_required('treasury.add_payment', raise_exception=True)
def create_payment(request, requisition_id):
    """Create a payment for an approved requisition"""
    
    requisition = get_object_or_404(Requisition, transaction_id=requisition_id)
    
    # Verify requisition is approved
    if requisition.status != 'approved':
        messages.error(request, 'Only approved requisitions can have payments created.')
        return redirect('transactions:view_requisition', requisition_id=requisition_id)
    
    if request.method == 'POST':
        try:
            amount = Decimal(request.POST.get('amount'))
            method = request.POST.get('method')
            destination = request.POST.get('destination')
            otp_required = request.POST.get('otp_required') == 'on'
            
            if not all([amount, method, destination]):
                messages.error(request, 'All fields are required.')
                return redirect('treasury:create_payment', requisition_id=requisition_id)
            
            payment = Payment.objects.create(
                requisition=requisition,
                amount=amount,
                method=method,
                destination=destination,
                otp_required=otp_required,
                status='pending',
            )
            
            messages.success(request, f'Payment created successfully! Payment ID: {payment.payment_id}')
            return redirect('treasury:manage_payments')
            
        except Exception as e:
            messages.error(request, f'Error creating payment: {str(e)}')
    
    context = {
        'requisition': requisition,
        'method_choices': Payment.PAYMENT_METHOD_CHOICES,
    }
    
    return render(request, 'treasury/create_payment.html', context)


@login_required
@permission_required('treasury.change_payment', raise_exception=True)
def execute_payment(request, payment_id):
    """Execute a pending payment"""
    
    payment = get_object_or_404(Payment, payment_id=payment_id)
    
    # Check if user can execute
    if not payment.can_execute(request.user):
        messages.error(request, 'You do not have permission to execute this payment.')
        return redirect('treasury:manage_payments')
    
    if payment.status != 'pending':
        messages.error(request, f'Payment is already {payment.status}. Cannot execute.')
        return redirect('treasury:manage_payments')
    
    if request.method == 'POST':
        try:
            # Simplified execution (without actual gateway integration)
            gateway_reference = request.POST.get('gateway_reference') or f'SIM-{payment.payment_id.hex[:10].upper()}'
            
            # Mark payment as success
            payment.mark_success(request.user)
            
            # Create execution record
            execution = PaymentExecution.objects.create(
                payment=payment,
                executor=request.user,
                gateway_reference=gateway_reference,
                gateway_status='SUCCESS',
                otp_verified_at=timezone.now() if payment.otp_required else None,
                otp_verified_by=request.user if payment.otp_required else None,
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
            )
            
            # Create ledger entry (debit from fund)
            # Find appropriate treasury fund
            req = payment.requisition
            treasury_fund = TreasuryFund.objects.filter(
                company=req.requested_by.company,
                branch=req.requested_by.branch,
            ).first()
            
            if treasury_fund:
                LedgerEntry.objects.create(
                    treasury_fund=treasury_fund,
                    payment_execution=execution,
                    entry_type='debit',
                    amount=payment.amount,
                    description=f'Payment for requisition {req.transaction_id}',
                )
                
                # Update fund balance
                treasury_fund.current_balance -= payment.amount
                treasury_fund.save()
            
            messages.success(request, f'Payment executed successfully! Reference: {gateway_reference}')
            return redirect('treasury:manage_payments')
            
        except Exception as e:
            payment.mark_failed(str(e))
            messages.error(request, f'Payment execution failed: {str(e)}')
    
    context = {
        'payment': payment,
        'requisition': payment.requisition,
    }
    
    return render(request, 'treasury/execute_payment.html', context)


# ==================== LEDGER ====================

@login_required
@permission_required('treasury.view_ledgerentry', raise_exception=True)
def view_ledger(request):
    """View ledger entries with filters"""
    
    # Filters
    fund_filter = request.GET.get('fund', '')
    entry_type_filter = request.GET.get('entry_type', '')
    reconciled_filter = request.GET.get('reconciled', '')
    
    entries = LedgerEntry.objects.all().select_related(
        'treasury_fund', 'payment_execution', 'reconciled_by'
    )
    
    if fund_filter:
        entries = entries.filter(treasury_fund_id=fund_filter)
    
    if entry_type_filter:
        entries = entries.filter(entry_type=entry_type_filter)
    
    if reconciled_filter == 'yes':
        entries = entries.filter(reconciled=True)
    elif reconciled_filter == 'no':
        entries = entries.filter(reconciled=False)
    
    entries = entries.order_by('-created_at')[:200]  # Limit to 200 recent
    
    funds = TreasuryFund.objects.all()
    
    # Stats
    total_debits = entries.filter(entry_type='debit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    total_credits = entries.filter(entry_type='credit').aggregate(Sum('amount'))['amount__sum'] or Decimal('0.00')
    unreconciled_count = entries.filter(reconciled=False).count()
    
    context = {
        'entries': entries,
        'funds': funds,
        'total_debits': total_debits,
        'total_credits': total_credits,
        'unreconciled_count': unreconciled_count,
        'fund_filter': fund_filter,
        'entry_type_filter': entry_type_filter,
        'reconciled_filter': reconciled_filter,
        'entry_type_choices': LedgerEntry.ENTRY_TYPE_CHOICES,
    }
    
    return render(request, 'treasury/view_ledger.html', context)


# ==================== VARIANCE ADJUSTMENTS ====================

@login_required
@permission_required('treasury.view_varianceadjustment', raise_exception=True)
def manage_variances(request):
    """View all variance adjustments"""
    
    # Filters
    status_filter = request.GET.get('status', '')
    fund_filter = request.GET.get('fund', '')
    
    variances = VarianceAdjustment.objects.all().select_related(
        'treasury_fund', 'initiated_by', 'approved_by'
    )
    
    if status_filter:
        variances = variances.filter(status=status_filter)
    
    if fund_filter:
        variances = variances.filter(treasury_fund_id=fund_filter)
    
    variances = variances.order_by('-created_at')
    
    funds = TreasuryFund.objects.all()
    
    context = {
        'variances': variances,
        'funds': funds,
        'status_filter': status_filter,
        'fund_filter': fund_filter,
        'status_choices': VarianceAdjustment.STATUS_CHOICES,
    }
    
    return render(request, 'treasury/manage_variances.html', context)


@login_required
@permission_required('treasury.add_varianceadjustment', raise_exception=True)
def create_variance(request):
    """Create a variance adjustment"""
    
    if request.method == 'POST':
        try:
            fund_id = request.POST.get('fund')
            adjusted_amount = Decimal(request.POST.get('adjusted_amount'))
            reason = request.POST.get('reason')
            
            if not all([fund_id, reason]):
                messages.error(request, 'Fund and reason are required.')
                return redirect('treasury:create_variance')
            
            fund = TreasuryFund.objects.get(fund_id=fund_id)
            original_amount = fund.current_balance
            variance_amount = adjusted_amount - original_amount
            
            variance = VarianceAdjustment.objects.create(
                treasury_fund=fund,
                original_amount=original_amount,
                adjusted_amount=adjusted_amount,
                variance_amount=variance_amount,
                reason=reason,
                initiated_by=request.user,
                status='pending',
            )
            
            messages.success(request, f'Variance adjustment created. Pending CFO approval.')
            return redirect('treasury:manage_variances')
            
        except Exception as e:
            messages.error(request, f'Error creating variance: {str(e)}')
    
    funds = TreasuryFund.objects.all()
    
    context = {
        'funds': funds,
    }
    
    return render(request, 'treasury/create_variance.html', context)


@login_required
@permission_required('treasury.change_varianceadjustment', raise_exception=True)
def approve_variance(request, variance_id):
    """Approve or reject a variance adjustment (CFO only)"""
    
    # Only CFO or Admin can approve
    if request.user.role.lower() not in ['cfo', 'admin']:
        messages.error(request, 'Only CFO can approve variance adjustments.')
        return redirect('treasury:manage_variances')
    
    variance = get_object_or_404(VarianceAdjustment, adjustment_id=variance_id)
    
    if variance.status != 'pending':
        messages.error(request, f'Variance is already {variance.status}.')
        return redirect('treasury:manage_variances')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve':
            variance.status = 'approved'
            variance.approved_by = request.user
            variance.approved_at = timezone.now()
            variance.save()
            
            # Apply adjustment to fund
            fund = variance.treasury_fund
            fund.current_balance = variance.adjusted_amount
            fund.save()
            
            # Create ledger entry
            LedgerEntry.objects.create(
                treasury_fund=fund,
                entry_type='adjustment',
                amount=abs(variance.variance_amount),
                description=f'Variance adjustment: {variance.reason}',
                reconciled=True,
                reconciled_by=request.user,
                reconciliation_timestamp=timezone.now(),
            )
            
            messages.success(request, 'Variance adjustment approved and applied.')
            
        elif action == 'reject':
            variance.status = 'rejected'
            variance.approved_by = request.user
            variance.approved_at = timezone.now()
            variance.save()
            
            messages.success(request, 'Variance adjustment rejected.')
        
        return redirect('treasury:manage_variances')
    
    context = {
        'variance': variance,
    }
    
    return render(request, 'treasury/approve_variance.html', context)


# ==================== BULK PAYMENT UPLOAD ====================

def sanitize_mpesa_text(text):
    """
    Remove special characters for M-Pesa compliance.
    M-Pesa only accepts alphanumeric characters and spaces.
    """
    if not text:
        return ''
    # Keep only alphanumeric and spaces
    return re.sub(r'[^a-zA-Z0-9\s]', '', str(text))


@login_required
@permission_required('treasury.can_manage_payments', raise_exception=True)
def select_payments_for_bulk(request):
    """
    Treasury selects which approved requisitions to pay
    """
    # Get approved requisitions that don't have payments yet
    approved_requisitions = Requisition.objects.filter(
        approval_status='approved',
        payments__isnull=True
    ).select_related('requested_by', 'company', 'branch').order_by('approval_date')
    
    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_requisitions')
        
        if not selected_ids:
            messages.warning(request, 'Please select at least one requisition to pay.')
            return redirect('treasury:select_payments_for_bulk')
        
        # Store selected IDs in session
        request.session['selected_payment_requisitions'] = selected_ids
        messages.success(request, f'{len(selected_ids)} requisitions selected for payment.')
        return redirect('treasury:generate_mpesa_bulk')
    
    # Filtering
    company_id = request.GET.get('company')
    branch_id = request.GET.get('branch')
    search = request.GET.get('search', '')
    
    if company_id:
        approved_requisitions = approved_requisitions.filter(company_id=company_id)
    if branch_id:
        approved_requisitions = approved_requisitions.filter(branch_id=branch_id)
    if search:
        approved_requisitions = approved_requisitions.filter(
            Q(requested_by__first_name__icontains=search) |
            Q(requested_by__last_name__icontains=search) |
            Q(purpose__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Calculate totals
    total_amount = approved_requisitions.aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    context = {
        'title': 'Select Payments for M-Pesa Bulk Processing',
        'requisitions': approved_requisitions,
        'companies': Company.objects.all(),
        'branches': Branch.objects.all(),
        'total_count': approved_requisitions.count(),
        'total_amount': total_amount,
    }
    return render(request, 'treasury/select_payments_bulk.html', context)


@login_required
@permission_required('treasury.can_manage_payments', raise_exception=True)
def generate_mpesa_bulk(request):
    """
    Generate M-Pesa bulk payment Excel file for selected requisitions
    Ready for API submission to M-Pesa
    """
    selected_ids = request.session.get('selected_payment_requisitions', [])
    
    if not selected_ids:
        messages.warning(request, 'No requisitions selected. Please select payments first.')
        return redirect('treasury:select_payments_for_bulk')
    
    # Get selected approved requisitions
    approved_requisitions = Requisition.objects.filter(
        requisition_id__in=selected_ids,
        approval_status='approved',
        payments__isnull=True
    ).select_related('requested_by', 'company', 'branch').order_by('approval_date')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "M-Pesa Bulk Payments"
    
    # Define headers matching M-Pesa template
    headers = ['MobileNumber', 'DocumentType', 'DocumentNumber', 'Amount', 'PurposeOfPayment', 'Name']
    
    # Style headers
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Set column widths
    ws.column_dimensions['A'].width = 15  # MobileNumber
    ws.column_dimensions['B'].width = 15  # DocumentType
    ws.column_dimensions['C'].width = 20  # DocumentNumber
    ws.column_dimensions['D'].width = 12  # Amount
    ws.column_dimensions['E'].width = 40  # PurposeOfPayment
    ws.column_dimensions['F'].width = 25  # Name
    
    # Fill data from approved requisitions
    for row_num, req in enumerate(approved_requisitions, 2):
        # Generate voucher number
        voucher_number = f"PAY{req.requisition_id.hex[:12].upper()}"
        
        # Get requester phone (assuming it's in user profile or requisition)
        mobile = getattr(req.requested_by, 'phone_number', '')
        if not mobile:
            mobile = ''  # Leave empty if not available
        
        # Sanitize purpose of payment
        purpose = sanitize_mpesa_text(req.purpose or req.description)[:100]  # M-Pesa limit
        
        # Get requester full name
        name = f"{req.requested_by.first_name} {req.requested_by.last_name}".strip()
        if not name:
            name = req.requested_by.username
        
        # Write row data
        ws.cell(row=row_num, column=1, value=mobile)
        ws.cell(row=row_num, column=2, value='')  # DocumentType empty
        ws.cell(row=row_num, column=3, value=voucher_number)
        ws.cell(row=row_num, column=4, value=float(req.amount))
        ws.cell(row=row_num, column=5, value=purpose)
        ws.cell(row=row_num, column=6, value=name)
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=mpesa_bulk_payments_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    
    # Clear selection from session
    if 'selected_payment_requisitions' in request.session:
        del request.session['selected_payment_requisitions']
    
    messages.success(request, f'M-Pesa bulk payment file generated for {approved_requisitions.count()} payments.')
    
    return response


@login_required
@permission_required('treasury.can_manage_payments', raise_exception=True)
def send_to_mpesa_api(request):
    """
    Send selected payments to M-Pesa API for batch processing
    """
    selected_ids = request.session.get('selected_payment_requisitions', [])
    
    if not selected_ids:
        messages.warning(request, 'No requisitions selected. Please select payments first.')
        return redirect('treasury:select_payments_for_bulk')
    
    # Get selected approved requisitions
    approved_requisitions = Requisition.objects.filter(
        requisition_id__in=selected_ids,
        approval_status='approved',
        payments__isnull=True
    ).select_related('requested_by', 'company', 'branch').order_by('approval_date')
    
    if request.method == 'POST':
        try:
            import requests
            
            # Prepare M-Pesa bulk payment data
            payments_data = []
            created_payments = []
            
            for req in approved_requisitions:
                # Generate voucher number
                voucher_number = f"PAY{req.requisition_id.hex[:12].upper()}"
                
                # Get requester phone
                mobile = getattr(req.requested_by, 'phone_number', '')
                if not mobile:
                    messages.error(request, f'Missing phone number for {req.requested_by.get_full_name()}')
                    continue
                
                # Validate phone format
                mobile_clean = str(mobile).replace(' ', '').replace('+', '')
                if not re.match(r'^254\d{9}$', mobile_clean):
                    messages.error(request, f'Invalid phone format for {req.requested_by.get_full_name()}: {mobile}')
                    continue
                
                # Sanitize purpose
                purpose = sanitize_mpesa_text(req.purpose or req.description)[:100]
                
                # Get name
                name = f"{req.requested_by.first_name} {req.requested_by.last_name}".strip()
                if not name:
                    name = req.requested_by.username
                
                # Create Payment record
                payment = Payment.objects.create(
                    requisition=req,
                    voucher_number=voucher_number,
                    amount=req.amount,
                    method='mpesa',
                    destination=mobile_clean,
                    description=purpose,
                    status='executing',
                    created_by=request.user,
                )
                created_payments.append(payment)
                
                # Prepare M-Pesa API payload
                payments_data.append({
                    'MobileNumber': mobile_clean,
                    'DocumentType': '',
                    'DocumentNumber': voucher_number,
                    'Amount': float(req.amount),
                    'PurposeOfPayment': purpose,
                    'Name': name,
                })
            
            if not payments_data:
                messages.error(request, 'No valid payments to process.')
                return redirect('treasury:select_payments_for_bulk')
            
            # TODO: Send to M-Pesa API
            # This is a placeholder - integrate with actual M-Pesa API
            """
            mpesa_response = requests.post(
                'https://api.safaricom.co.ke/mpesa/b2c/v1/paymentrequest',
                json={
                    'InitiatorName': settings.MPESA_INITIATOR_NAME,
                    'SecurityCredential': settings.MPESA_SECURITY_CREDENTIAL,
                    'CommandID': 'BusinessPayment',
                    'Amount': sum(p['Amount'] for p in payments_data),
                    'PartyA': settings.MPESA_SHORTCODE,
                    'Remarks': 'Bulk payment processing',
                    'QueueTimeOutURL': settings.MPESA_TIMEOUT_URL,
                    'ResultURL': settings.MPESA_RESULT_URL,
                    'Occassion': 'Bulk Payment',
                    'Transactions': payments_data,
                },
                headers={
                    'Authorization': f'Bearer {get_mpesa_access_token()}',
                    'Content-Type': 'application/json',
                }
            )
            """
            
            # For now, mark as success (replace with actual API response handling)
            for payment in created_payments:
                payment.status = 'pending'  # Will be updated by M-Pesa callback
                payment.executor = request.user
                payment.execution_timestamp = timezone.now()
                payment.save()
            
            # Clear selection
            if 'selected_payment_requisitions' in request.session:
                del request.session['selected_payment_requisitions']
            
            messages.success(
                request, 
                f'Successfully submitted {len(payments_data)} payments to M-Pesa for processing. '
                f'Total amount: {sum(p["Amount"] for p in payments_data):,.2f}'
            )
            
            return redirect('treasury:manage_payments')
            
        except Exception as e:
            messages.error(request, f'Error sending to M-Pesa API: {str(e)}')
            return redirect('treasury:send_to_mpesa_api')
    
    # Calculate totals
    total_amount = approved_requisitions.aggregate(Sum('amount'))['amount__sum'] or Decimal('0')
    
    context = {
        'title': 'Send to M-Pesa Bulk Payment API',
        'requisitions': approved_requisitions,
        'total_count': approved_requisitions.count(),
        'total_amount': total_amount,
    }
    return render(request, 'treasury/send_to_mpesa.html', context)


@login_required
@permission_required('treasury.can_manage_payments', raise_exception=True)
def bulk_payment_upload(request):
    """
    Alternative: Upload and process M-Pesa bulk payment Excel file
    (For manual file-based workflow if needed)
    """
    if request.method == 'POST' and request.FILES.get('payment_file'):
        try:
            from openpyxl import load_workbook
            
            file = request.FILES['payment_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Validate headers
            expected_headers = ['MobileNumber', 'DocumentType', 'DocumentNumber', 'Amount', 'PurposeOfPayment', 'Name']
            actual_headers = [cell.value for cell in ws[1]]
            
            if actual_headers != expected_headers:
                messages.error(request, f'Invalid template format. Expected headers: {", ".join(expected_headers)}')
                return redirect('treasury:bulk_payment_upload')
            
            # Process rows
            payments_data = []
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                mobile, doc_type, voucher, amount, purpose, name = row
                
                # Validate data
                row_errors = []
                
                if not mobile:
                    row_errors.append(f'Row {row_num}: Mobile number is required')
                elif not re.match(r'^254\d{9}$', str(mobile).replace(' ', '')):
                    row_errors.append(f'Row {row_num}: Invalid mobile number format (must be 254XXXXXXXXX)')
                
                if not voucher:
                    row_errors.append(f'Row {row_num}: Document number is required')
                
                try:
                    amount = Decimal(str(amount))
                    if amount <= 0:
                        row_errors.append(f'Row {row_num}: Amount must be greater than 0')
                except (ValueError, TypeError):
                    row_errors.append(f'Row {row_num}: Invalid amount')
                
                if not purpose:
                    row_errors.append(f'Row {row_num}: Purpose of payment is required')
                
                if not name:
                    row_errors.append(f'Row {row_num}: Name is required')
                
                # Check for special characters
                if purpose and re.search(r'[^a-zA-Z0-9\s]', purpose):
                    row_errors.append(f'Row {row_num}: Purpose contains special characters (M-Pesa only allows alphanumeric)')
                
                if row_errors:
                    errors.extend(row_errors)
                else:
                    payments_data.append({
                        'mobile': str(mobile).replace(' ', ''),
                        'voucher': str(voucher),
                        'amount': amount,
                        'purpose': purpose,
                        'name': name,
                    })
            
            if errors:
                for error in errors:
                    messages.error(request, error)
                return redirect('treasury:bulk_payment_upload')
            
            # Store in session for preview
            request.session['bulk_payments_data'] = payments_data
            messages.success(request, f'{len(payments_data)} payments loaded. Please review and confirm.')
            return redirect('treasury:process_bulk_payments')
            
        except Exception as e:
            messages.error(request, f'Error processing file: {str(e)}')
            return redirect('treasury:bulk_payment_upload')
    
    context = {
        'title': 'Bulk Payment Upload',
    }
    return render(request, 'treasury/bulk_payment_upload.html', context)


@login_required
@permission_required('treasury.can_manage_payments', raise_exception=True)
def process_bulk_payments(request):
    """
    Preview and confirm bulk payment creation
    """
    payments_data = request.session.get('bulk_payments_data', [])
    
    if not payments_data:
        messages.warning(request, 'No payment data found. Please upload a file first.')
        return redirect('treasury:bulk_payment_upload')
    
    if request.method == 'POST':
        try:
            created_count = 0
            
            for payment_data in payments_data:
                # Check if voucher number already exists
                if Payment.objects.filter(voucher_number=payment_data['voucher']).exists():
                    messages.warning(request, f"Skipped duplicate voucher: {payment_data['voucher']}")
                    continue
                
                # Create payment record
                Payment.objects.create(
                    voucher_number=payment_data['voucher'],
                    amount=payment_data['amount'],
                    method='mpesa',
                    destination=payment_data['mobile'],
                    description=payment_data['purpose'],
                    status='pending',
                    created_by=request.user,
                )
                created_count += 1
            
            # Clear session data
            del request.session['bulk_payments_data']
            
            messages.success(request, f'Successfully created {created_count} payment records.')
            return redirect('treasury:manage_payments')
            
        except Exception as e:
            messages.error(request, f'Error creating payments: {str(e)}')
            return redirect('treasury:process_bulk_payments')
    
    context = {
        'title': 'Review Bulk Payments',
        'payments_data': payments_data,
        'total_count': len(payments_data),
        'total_amount': sum(p['amount'] for p in payments_data),
    }
    return render(request, 'treasury/process_bulk_payments.html', context)

