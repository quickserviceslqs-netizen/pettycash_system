"""
Bulk Import for Organization Entities
Handles Excel uploads for Companies, Regions, Branches, Departments, Cost Centers, Positions
"""

import csv
import io

import openpyxl
import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from openpyxl.styles import Font, PatternFill

from organization.models import (
    Branch,
    Company,
    CostCenter,
    Department,
    Position,
    Region,
)
from settings_manager.views import log_activity

# ============================================
# COMPANIES
# ============================================


@login_required
@permission_required("organization.add_company", raise_exception=True)
def download_companies_template(request):
    """Download Excel template for bulk company import"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Companies"

    # Headers
    headers = ["name", "code"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )

    # Single example row
    ws.append(["DemoCo", "DEMO"])

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    ws2["A1"] = "Company Import Instructions"
    ws2["A1"].font = Font(bold=True, size=14)
    instructions = [
        "",
        "FIELD REQUIREMENTS:",
        "• name: Company name (must be unique)",
        "• code: Short code (2-10 characters, must be unique)",
        "",
        "INSTRUCTIONS:",
        "1. Fill in the 'Companies' sheet with your data",
        "2. Delete the example rows before uploading",
        "3. Save and upload the file",
    ]
    for row_num, instruction in enumerate(instructions, 2):
        ws2[f"A{row_num}"] = instruction
    ws2.column_dimensions["A"].width = 60

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        'attachment; filename="companies_import_template.xlsx"'
    )
    wb.save(response)
    return response


@login_required
@permission_required("organization.add_company", raise_exception=True)
def import_companies(request):
    """Bulk import companies from Excel or CSV"""
    if request.method == "POST" and request.FILES.get("csv_file"):
        uploaded_file = request.FILES["csv_file"]

        is_excel = uploaded_file.name.endswith((".xlsx", ".xls"))
        is_csv = uploaded_file.name.endswith(".csv")

        if not (is_excel or is_csv):
            messages.error(request, "Please upload an Excel or CSV file")
            return redirect("import_organizations")

        try:
            if is_excel:
                df = pd.read_excel(uploaded_file)
                data_rows = df.to_dict("records")
            else:
                decoded_file = uploaded_file.read().decode("utf-8")
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                data_rows = list(reader)

            success_count = 0
            error_count = 0
            errors = []

            with transaction.atomic():
                for row_num, row in enumerate(data_rows, start=2):
                    try:
                        if is_excel:
                            row = {
                                k: (str(v) if pd.notna(v) else "")
                                for k, v in row.items()
                            }

                        if not row.get("name") or str(row.get("name", "")).startswith(
                            "INSTRUCTIONS"
                        ):
                            continue

                        name = str(row["name"]).strip()
                        code = str(row["code"]).strip()

                        if not name or not code:
                            errors.append(f"Row {row_num}: Missing name or code")
                            error_count += 1
                            continue

                        if Company.objects.filter(name__iexact=name).exists():
                            errors.append(
                                f"Row {row_num}: Company '{name}' already exists"
                            )
                            error_count += 1
                            continue

                        if Company.objects.filter(code__iexact=code).exists():
                            errors.append(
                                f"Row {row_num}: Code '{code}' already exists"
                            )
                            error_count += 1
                            continue

                        Company.objects.create(name=name, code=code)
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1

            if success_count > 0:
                messages.success(
                    request,
                    f'✅ Successfully imported {success_count} compan{"y" if success_count == 1 else "ies"}',
                )
                log_activity(
                    request.user,
                    "ORG_BULK_IMPORT",
                    f"Imported {success_count} companies",
                )

            if error_count > 0:
                messages.warning(request, f"⚠️ {error_count} row(s) failed")
                for error in errors[:10]:
                    messages.error(request, error)

            return redirect("import_organizations")

        except Exception as e:
            messages.error(request, f"Error processing CSV: {str(e)}")
            return redirect("import_organizations")

    return redirect("import_organizations")


# ============================================
# REGIONS
# ============================================


@login_required
@permission_required("organization.add_region", raise_exception=True)
def download_regions_template(request):
    """Download CSV template for bulk region import"""
    # Get filter parameter
    company_id = request.GET.get("company")

    response = HttpResponse(content_type="text/csv")

    # Filter companies if specified
    if company_id:
        companies = Company.objects.filter(id=company_id)
        company_name = companies.first().name if companies.exists() else "filtered"
        response["Content-Disposition"] = (
            f'attachment; filename="regions_{company_name.replace(" ", "_")}_template.csv"'
        )
    else:
        companies = Company.objects.all()
        response["Content-Disposition"] = (
            'attachment; filename="regions_import_template.csv"'
        )

    writer = csv.writer(response)
    writer.writerow(["name", "code", "company_name"])
    # Single example row
    if companies.exists():
        example_company = companies.first()
        writer.writerow(
            ["DemoRegion", f"{example_company.code}DR", example_company.name]
        )
    else:
        writer.writerow(["DemoRegion", "DR", "DemoCo"])

    writer.writerow([])
    writer.writerow(["INSTRUCTIONS:"])
    writer.writerow(["- name: Region name"])
    writer.writerow(["- code: Short code (2-10 characters, must be unique)"])
    writer.writerow(["- company_name: Must exactly match existing company name"])
    if company_id:
        writer.writerow([f"- FILTERED: Only showing regions for selected company"])
    writer.writerow([])
    writer.writerow(["AVAILABLE COMPANIES: (omitted for brevity)"])

    return response


@login_required
@permission_required("organization.add_region", raise_exception=True)
def import_regions(request):
    """Bulk import regions from CSV"""
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]
        is_excel = csv_file.name.endswith((".xlsx", ".xls"))
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a CSV file")
            return redirect("import_organizations")
        try:
            decoded_file = csv_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            data_rows = list(reader)
            success_count = 0
            error_count = 0
            errors = []
            with transaction.atomic():
                for row_num, row in enumerate(data_rows, start=2):
                    try:
                        if is_excel:
                            row = {
                                k: (str(v) if pd.notna(v) else "")
                                for k, v in row.items()
                            }
                        if not row.get("name") or str(row.get("name", "")).startswith(
                            "INSTRUCTIONS"
                        ):
                            continue
                        name = str(row["name"]).strip()
                        code = str(row["code"]).strip()
                        company_name = str(row["company_name"]).strip()
                        if not all([name, code, company_name]):
                            errors.append(f"Row {row_num}: Missing required fields")
                            error_count += 1
                            continue

                        try:
                            company = Company.objects.get(name__iexact=company_name)
                        except Company.DoesNotExist:
                            errors.append(
                                f"Row {row_num}: Company '{company_name}' not found"
                            )
                            error_count += 1
                            continue

                        if Region.objects.filter(code__iexact=code).exists():
                            errors.append(
                                f"Row {row_num}: Code '{code}' already exists"
                            )
                            error_count += 1
                            continue

                        Region.objects.create(name=name, code=code, company=company)
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1

            if success_count > 0:
                messages.success(
                    request, f"✅ Successfully imported {success_count} region(s)"
                )
                log_activity(
                    request.user, "ORG_BULK_IMPORT", f"Imported {success_count} regions"
                )

            if error_count > 0:
                messages.warning(request, f"⚠️ {error_count} row(s) failed")
                for error in errors[:10]:
                    messages.error(request, error)

            return redirect("import_organizations")

        except Exception as e:
            messages.error(request, f"Error processing CSV: {str(e)}")
            return redirect("import_organizations")

    return redirect("import_organizations")


# ============================================
# BRANCHES
# ============================================


@login_required
@permission_required("organization.add_branch", raise_exception=True)
def download_branches_template(request):
    """Download CSV template for bulk branch import"""
    # Get filter parameters
    company_id = request.GET.get("company")
    region_id = request.GET.get("region")

    response = HttpResponse(content_type="text/csv")

    # Filter regions based on parameters
    regions = Region.objects.select_related("company").all()
    filename_parts = ["branches"]

    if region_id:
        regions = regions.filter(id=region_id)
        if regions.exists():
            filename_parts.append(regions.first().name.replace(" ", "_"))
    elif company_id:
        regions = regions.filter(company_id=company_id)
        company = Company.objects.filter(id=company_id).first()
        if company:
            filename_parts.append(company.name.replace(" ", "_"))

    response["Content-Disposition"] = (
        f'attachment; filename="{"_".join(filename_parts)}_template.csv"'
    )

    writer = csv.writer(response)
    writer.writerow(["name", "code", "phone", "region_name", "company_name"])
    # Single example row
    if regions.exists():
        r = regions.first()
        writer.writerow(
            ["DemoBranch", f"{r.code}DB", "+254700000000", r.name, r.company.name]
        )
    else:
        writer.writerow(["DemoBranch", "DB", "+254700000000", "DemoRegion", "DemoCo"])

    writer.writerow([])
    writer.writerow(["INSTRUCTIONS:"])
    writer.writerow(["- name: Branch name (must be unique)"])
    writer.writerow(["- code: Short code (2-10 characters, must be unique)"])
    writer.writerow(["- phone: Contact phone number (optional)"])
    writer.writerow(["- region_name: Must exactly match existing region"])
    writer.writerow(["- company_name: Must exactly match existing company"])
    if region_id:
        writer.writerow([f"- FILTERED: Only showing branches for selected region"])
    elif company_id:
        writer.writerow([f"- FILTERED: Only showing branches for selected company"])
    writer.writerow([])
    writer.writerow(["AVAILABLE REGIONS (omitted for brevity)"])

    return response


@login_required
@permission_required("organization.add_branch", raise_exception=True)
def import_branches(request):
    """Bulk import branches from CSV"""
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]
        is_excel = csv_file.name.endswith((".xlsx", ".xls"))
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a CSV file")
            return redirect("import_organizations")
        try:
            decoded_file = csv_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            data_rows = list(reader)
            success_count = 0
            error_count = 0
            errors = []
            with transaction.atomic():
                for row_num, row in enumerate(data_rows, start=2):
                    try:
                        if is_excel:
                            row = {
                                k: (str(v) if pd.notna(v) else "")
                                for k, v in row.items()
                            }
                        if not row.get("name") or str(row.get("name", "")).startswith(
                            "INSTRUCTIONS"
                        ):
                            continue
                        name = str(row["name"]).strip()
                        code = str(row["code"]).strip()
                        phone = row.get("phone", "").strip()
                        region_name = str(row["region_name"]).strip()
                        company_name = str(row["company_name"]).strip()

                        if not all([name, code, region_name, company_name]):
                            errors.append(f"Row {row_num}: Missing required fields")
                            error_count += 1
                            continue

                        try:
                            company = Company.objects.get(name__iexact=company_name)
                        except Company.DoesNotExist:
                            errors.append(
                                f"Row {row_num}: Company '{company_name}' not found"
                            )
                            error_count += 1
                            continue

                        try:
                            region = Region.objects.get(
                                name__iexact=region_name, company=company
                            )
                        except Region.DoesNotExist:
                            errors.append(
                                f"Row {row_num}: Region '{region_name}' not found for company '{company_name}'"
                            )
                            error_count += 1
                            continue

                        if Branch.objects.filter(name__iexact=name).exists():
                            errors.append(
                                f"Row {row_num}: Branch '{name}' already exists"
                            )
                            error_count += 1
                            continue

                        if Branch.objects.filter(code__iexact=code).exists():
                            errors.append(
                                f"Row {row_num}: Code '{code}' already exists"
                            )
                            error_count += 1
                            continue

                        Branch.objects.create(
                            name=name,
                            code=code,
                            phone=phone if phone else None,
                            region=region,
                        )
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1

            if success_count > 0:
                messages.success(
                    request, f"✅ Successfully imported {success_count} branch(es)"
                )
                log_activity(
                    request.user,
                    "ORG_BULK_IMPORT",
                    f"Imported {success_count} branches",
                )

            if error_count > 0:
                messages.warning(request, f"⚠️ {error_count} row(s) failed")
                for error in errors[:10]:
                    messages.error(request, error)

            return redirect("import_organizations")

        except Exception as e:
            messages.error(request, f"Error processing CSV: {str(e)}")
            return redirect("import_organizations")

    return redirect("import_organizations")


# ============================================
# DEPARTMENTS
# ============================================


@login_required
@permission_required("organization.add_department", raise_exception=True)
def download_departments_template(request):
    """Download CSV template for bulk department import"""
    # Get filter parameters
    company_id = request.GET.get("company")
    region_id = request.GET.get("region")
    branch_id = request.GET.get("branch")

    response = HttpResponse(content_type="text/csv")

    # Filter branches based on parameters
    branches = Branch.objects.select_related("region__company").all()
    filename_parts = ["departments"]

    if branch_id:
        branches = branches.filter(id=branch_id)
        if branches.exists():
            filename_parts.append(branches.first().name.replace(" ", "_"))
    elif region_id:
        branches = branches.filter(region_id=region_id)
        region = Region.objects.filter(id=region_id).first()
        if region:
            filename_parts.append(region.name.replace(" ", "_"))
    elif company_id:
        branches = branches.filter(region__company_id=company_id)
        company = Company.objects.filter(id=company_id).first()
        if company:
            filename_parts.append(company.name.replace(" ", "_"))

    response["Content-Disposition"] = (
        f'attachment; filename="{"_".join(filename_parts)}_template.csv"'
    )

    writer = csv.writer(response)
    writer.writerow(["name", "branch_name"])
    # Single example row
    if branches.exists():
        b = branches.first()
        writer.writerow(["DemoDepartment", b.name])
    else:
        writer.writerow(["DemoDepartment", "DemoBranch"])

    writer.writerow([])
    writer.writerow(["INSTRUCTIONS:"])
    writer.writerow(["- name: Department name"])
    writer.writerow(["- branch_name: Must exactly match existing branch"])
    writer.writerow(["- Same department name can exist in multiple branches"])
    if branch_id:
        writer.writerow([f"- FILTERED: Only showing departments for selected branch"])
    elif region_id:
        writer.writerow([f"- FILTERED: Only showing departments for selected region"])
    elif company_id:
        writer.writerow([f"- FILTERED: Only showing departments for selected company"])
    writer.writerow([])
    writer.writerow(["AVAILABLE BRANCHES (omitted for brevity)"])

    return response


@login_required
@permission_required("organization.add_department", raise_exception=True)
def import_departments(request):
    """Bulk import departments from CSV"""
    if request.method == "POST" and request.FILES.get("csv_file"):
        csv_file = request.FILES["csv_file"]
        is_excel = csv_file.name.endswith((".xlsx", ".xls"))
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a CSV file")
            return redirect("import_organizations")
        try:
            decoded_file = csv_file.read().decode("utf-8")
            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)
            data_rows = list(reader)
            success_count = 0
            error_count = 0
            errors = []
            with transaction.atomic():
                for row_num, row in enumerate(data_rows, start=2):
                    try:
                        if is_excel:
                            row = {
                                k: (str(v) if pd.notna(v) else "")
                                for k, v in row.items()
                            }
                        if not row.get("name") or str(row.get("name", "")).startswith(
                            "INSTRUCTIONS"
                        ):
                            continue
                        name = str(row["name"]).strip()
                        branch_name = str(row["branch_name"]).strip()
                        if not all([name, branch_name]):
                            errors.append(f"Row {row_num}: Missing required fields")
                            error_count += 1
                            continue

                        try:
                            branch = Branch.objects.get(name__iexact=branch_name)
                        except Branch.DoesNotExist:
                            errors.append(
                                f"Row {row_num}: Branch '{branch_name}' not found"
                            )
                            error_count += 1
                            continue

                        if Department.objects.filter(
                            name__iexact=name, branch=branch
                        ).exists():
                            errors.append(
                                f"Row {row_num}: Department '{name}' already exists in branch '{branch_name}'"
                            )
                            error_count += 1
                            continue

                        Department.objects.create(name=name, branch=branch)
                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {row_num}: {str(e)}")
                        error_count += 1

            if success_count > 0:
                messages.success(
                    request, f"✅ Successfully imported {success_count} department(s)"
                )
                log_activity(
                    request.user,
                    "ORG_BULK_IMPORT",
                    f"Imported {success_count} departments",
                )

            if error_count > 0:
                messages.warning(request, f"⚠️ {error_count} row(s) failed")
                for error in errors[:10]:
                    messages.error(request, error)

            return redirect("import_organizations")

        except Exception as e:
            messages.error(request, f"Error processing CSV: {str(e)}")
            return redirect("import_organizations")

    return redirect("import_organizations")


# ============================================
# MAIN IMPORT PAGE
# ============================================


@login_required
def import_organizations(request):
    """Main page for bulk importing all organization entities"""
    context = {
        "companies_count": Company.objects.count(),
        "regions_count": Region.objects.count(),
        "branches_count": Branch.objects.count(),
        "departments_count": Department.objects.count(),
        "companies": Company.objects.all().order_by("name"),
        "regions": Region.objects.select_related("company")
        .all()
        .order_by("company__name", "name"),
        "branches": Branch.objects.select_related("region__company")
        .all()
        .order_by("region__company__name", "region__name", "name"),
    }
    return render(request, "organization/bulk_import.html", context)
